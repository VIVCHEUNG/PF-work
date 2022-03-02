import os
from re import sub
import pandas as pd
from rich.progress import track
import warnings
from rich.console import Console
from rich.markdown import Markdown
from rich.table import Table
from rich.padding import Padding
from rich import print
from rich.tree import Tree
from rich.layout import Layout
from rich import print
from rich.panel import Panel
import time
from rich.live import Live
from rich.align import Align
import sys
from rich.table import Table
from contextlib import contextmanager
from rich import box
from rich.text import Text
import openpyxl
import numpy as np
from datetime import datetime

warnings.filterwarnings('ignore')
console = Console()
layout = Layout()
console.log("Hello! Da ba dee da ba di~")


try:
    os.system('pip install visidata==2.4')
    os.system('pip install fabulous')
except:
    pass


def restart_program():
    python = sys.executable
    os.execl(python, python, *sys.argv)


tree0 = Tree("[red]日常业务 [#BEBEBE]（/投资资产核算管理/01 交易清算）")
tree1 = tree0.add("[purple]交易清算")
tree1_1 = tree1.add("40 每日交易清算")
tree1_1.add('德邦')
tree1_1.add('申万')
tree1_1.add('招商')

tree2 = tree0.add("[blue]资金余额/资金明细")
tree2_1 = tree2.add("10 每日头寸")
tree2_1.add('工行')
tree2_1.add('建行')
tree2_1.add('农行')
tree2_1.add('浦发')
tree2_1.add('招商')

tree3 = tree0.add("[yellow]投资款日结报告")
tree3_1 = tree3.add('50 投资款日结报告')
tree3_1.add('投资款日结报告')
tree3_1.add('分红型产品资产划拨日结报表')
tree3_1.add('策略型产品资产划拨日结报表')

treeI = Tree("[green]托管行估值表 [#BEBEBE]（/投资资产核算管理/03 估值管理）")
treeI.add('汇总')
treeI.add('德邦1,7,8号')
treeI.add('工商亚洲-QDII')
treeI.add('工行')
treeI.add('建行')
treeI.add('农行')
treeI.add('浦发')
treeI.add('申万10,16号')

treeII = Tree('[red]资产持仓查询报表[#BEBEBE]（/投资资产核算管理/06 财务输出报表）')
treeII.add('可供出售金融资产——成本[150301]')
treeII.add('交易性金融资产——成本[110101]')
treeII.add('持有至到期投资-成本[150101]')
treeII.add('贷款和应收款项-成本[130301]')
treeII.add('长期股权投资-成本[151101]')
treeII.add('其他应付应收款-成本[1133]')
treeII.add('其他货币资金-成本[150301]')
treeII.add('法定存款-成本[100201+154101]')
treeII.add('卖出回购金融资产-成本[211101]')
treeII.add('买入返售金融资产-成本[111101]')

start_tree0 = Align(Padding(tree0, (2, 18), expand=False), align='center')
start_tree1 = Align(Padding(treeI, (2, 18), expand=False), align='center')
start_tree2 = Align(Padding(treeII, (2, 18), expand=False), align='center')
start_pic1 = Align(Padding(
    '[blink] o                 /\' ) \n                      /\'   (                          ,\n                __/\'     )                        .\' `;\n   o      _.-~~~~\'          ``---..__             .\'   ;\n     _.--\'  b)                       ``--...____.\'   .\'\n    (     _.      )).      `-._                     <\n     `\|\|\|\|)-.....___.-     `-.         __...--\'-.\'.\n       `---......____...---`.___.\'----... .\'         `.;\n                                        `-`           `\'',
    (2, 18)), align='center')

'''
layout0 = Layout()
layout0.split_column(Layout(name="up"),Layout(name="low"))
layout0['up'].split_row(Layout(Panel(md0,border_style='red'),name='poet'),Layout(start_pic1,name="up-2"))
layout0['low'].split_row(Layout(start_tree0,name="每日作业"),Layout(start_tree1,name="月度作业"))
layout0['月度作业'].split_column(Layout(start_tree1,name="估值表合并"),Layout(start_tree2,name="资产持仓报表/资产收益报表"))
layout0['每日作业'].size=70
layout0['月度作业'].size=70
#layout0['估值表合并'].size=12

print(layout0)
'''
print(start_pic1, start_tree0, start_tree1, start_tree2)

# ----------------------------------------------------------------------------------------------------------------------

Action_confirm = console.input("[blink][bold magenta]要进行什么操作？")

if Action_confirm == '资金余额':
    Base_route = '//Volumes//fileshare//投资资产核算管理//01 交易清算//10 每日头寸'

    period = input('查看哪天的资金余额？（yyyymmdd）:')


    ## period=period.replace('-','')
    def alter_extension(old, new, path):
        for root, dirs, files in os.walk(path):
            print(root + '已转化格式')
            for i in files:
                if i.split('.')[1] == old:
                    os.rename(path + '/' + i, path + '/' + i.split('.')[0] + '.' + new)


    # for i in ['人保','华泰万能','华泰资本金','太保','工行0752','申万10号','申万16号','资本金外币',"QDII"]:
    #     try:
    #         alter_extension('xls','html','工行/'+i)
    #     except:
    #         pass

    # 工行
    def 工行_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('_')[
                        2]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_html(path + "/" + i, flavor='lxml')
                    balance = file[3][file[3].iloc[:, 1].str.contains("期末余额", na=False)].iloc[[0], [4]].values[0][0]
                    exec(f'{var_name}=balance')
                    exec(f'df_list.append({var_name})')
        df_list[0] = df_list[0].replace(",", "")
        return float(df_list[0])


    for i in ['人保', '华泰万能', '华泰资本金', '太保', '工行0752', '申万10号', '资本金外币']:
        exec(f"{i}=工行_to_df(Base_route+'/工行/{i}',['xls','{period}'])")


    # 农行
    def 农行_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('_')[2].replace('-',
                                                                             '')  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_excel(path + "/" + i)
                    balance = file.iloc[[2], [6]].values[0][0]
                    exec(f'{var_name}=balance')
                    exec(f'df_list.append({var_name})')
        df_list[0] = df_list[0].replace(",", "")
        return float(df_list[0])


    for i in ['德邦7号', '德邦8号', '农行7465资本金', '农行7473万能', '农行7481分红']:
        # print(Base_route + "/农行/%s" % i)
        # print(os.path.exists(Base_route + "/农行/%s" % i))
        abc_period = datetime.strptime(period, "%Y%m%d").strftime("%Y-%m-%d")
        exec(f"{i}=农行_to_df(Base_route+'/农行/{i}',['xls','{abc_period}'])")

    # QDII
    QDII_Consolidated = pd.read_html(Base_route + "/工行/QDII/Consolidated Cash Report_0000058_" + period + ".xls",header=0, skiprows=0)
    for subfile in os.listdir(Base_route + "/工行/QDII/"):
        new_base_route = Base_route + "/工行/QDII/"
        if not subfile.endswith('xls'):
            continue
        split_filename = subfile.split("_")
        # print(os.path.join(new_base_route, subfile))
        if split_filename[1].endswith('1391') and split_filename[2].split('.')[0] == period:
            QDII_1373 = pd.read_html(os.path.join(new_base_route, subfile), header=0, skiprows=0)
        elif split_filename[1].endswith('4503') and split_filename[2].split('.')[0] == period:
            QDII_4503 = pd.read_html(os.path.join(new_base_route, subfile), header=0, skiprows=0)
        elif split_filename[1].endswith('3275') and split_filename[2].split('.')[0] == period:
            QDII_3019 = pd.read_html(os.path.join(new_base_route, subfile), header=0, skiprows=0)
    # QDII_3019 = pd.read_excel(Base_route+"/工行/QDII/【复星保德信人寿保险QDII产品-自有资金】T-1日境内账户头寸变动表-HKD" + period + ".xls",header=0,skiprows=0)
    # QDII_4503 = pd.read_excel(Base_route+"/工行/QDII/【复星保德信人寿保险QDII产品-自有资金】T-1日境内账户头寸变动表-RMB" + period + ".xls",header=0,skiprows=0)
    # QDII_1373 = pd.read_excel(Base_route+"/工行/QDII/【复星保德信人寿保险QDII产品-自有资金】T-1日境内账户头寸变动表-USD" + period + ".xls",header=0,skiprows=0)

    QDII3019 = QDII_3019[7][QDII_3019[7].iloc[:, 1].str.contains("Ending Balance", na=False)].iloc[[0], [4]].values[0][0]
    QDII4503 = QDII_4503[7][QDII_4503[7].iloc[:, 1].str.contains("Ending Balance", na=False)].iloc[[0], [4]].values[0][0]
    QDII1373 = QDII_1373[7][QDII_1373[7].iloc[:, 1].str.contains("Ending Balance", na=False)].iloc[[0], [4]].values[0][0]
    QDII1104 = QDII_Consolidated[3][QDII_Consolidated[3].iloc[:, 2].str.contains("Closing Balance", na=False)].iloc[[1], [6]].values[0][0]
    QDII1922 = QDII_Consolidated[3][QDII_Consolidated[3].iloc[:, 2].str.contains("Closing Balance", na=False)].iloc[[0], [6]].values[0][0]


    # 建行
    def 建行_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('-')[
                        0]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_excel(path + "/" + i, skiprows=2).fillna(0)
                    balance = file[file.iloc[:, 0].str.contains("活期存款账户", na=False)].iloc[[0], [10]].values[0][0]
                    exec(f'{var_name}=balance')
                    exec(f'df_list.append({var_name})')
        ## df_list[0]=df_list[0].replace(",","")
        return float(df_list[0])


    建行 = 建行_to_df(Base_route + '/建行', ['xls', period])


    # 浦发
    def 浦发_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('_')[
                        3]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_excel(path + "/" + i, skiprows=0, dtype=str)
                    balance = file.iloc[[10], [7]].values[0][0]
                    exec(f'{var_name}=balance')
                    exec(f'df_list.append({var_name})')
        return float(df_list[0])


    浦发 = 浦发_to_df(Base_route + '/浦发', ['xls', period, '每日基金明细表'])


    # 招商
    def 招商_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('-')[
                        3]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_excel(path + "/" + i, skiprows=0)
                    balance = file[file.iloc[:, 0].str.contains("可用资金：", na=False)].iloc[[0], [1]].values[0][0]
                    exec(f'{var_name}=balance')
                    exec(f'df_list.append({var_name})')
        df_list[0] = df_list[0].replace(",", "")
        return float(df_list[0])


    招商 = 招商_to_df(Base_route + '/招商', ['xlsx', period, '普通账户对账单'])

    # 导出excel
    table = openpyxl.load_workbook(Base_route + "/△ 资金余额/模板（勿删）.xlsx")
    sheet = table.active
    total = ["太保", "华泰资本金", "工行0752", "人保", "华泰万能", "资本金外币", "QDII1104", "QDII1922", "QDII1373", "QDII4503", "QDII3019",
             "申万10号", "请手动填入申万16数据", "招商", "德邦7号", "德邦8号", "农行7465资本金", "农行7473万能", "农行7481分红", "浦发", "建行"]
    for j in range(2, 23):
        sheet.cell(row=j, column=7).value = total[j - 2]
    table.save(Base_route + "/△ 资金余额/" + period + "_头寸余额.xlsx")
    print("[bold magenta]资金余额表已生成,请前往「△ 资金余额」文件夹查看 [green]:heavy_check_mark:")

    restart_program()


# --------------------------------------------------------------------------------------------------------------------

elif Action_confirm == '资金明细':
    Base_route = '//Volumes//fileshare//投资资产核算管理//01 交易清算//10 每日头寸'
    os.system('clear')

    # 一首小诗
    try:
        Poet = Base_route + '/Poet.md'
        with open(Poet) as MARKDOWN:
            md = Markdown(MARKDOWN.read())
        console.print(md, justify="center")
    except:
        pass


    def 工行_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('_')[2]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_html(path + "/" + i, flavor='lxml')
                    file[3].drop(index=[0], columns=[0], inplace=True)
                    file[3]['交易日期'] = file[0].iloc[1, 0].replace(' ', '').replace('日期：', '').replace('年', '').replace('月', '').replace('日', '')
                    exec(f'{var_name}=file[3]')
                    exec(f'df_list.append({var_name})')
        return pd.concat(df_list).reset_index(drop=True)


    def 建行_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('-')[
                        0]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_excel(path + "/" + i, skiprows=2).fillna(0)
                    date = file.iloc[0, 0].replace('日期：', '').replace('-', '')
                    file['交易日期'] = date
                    file.drop(index=[0, 1, 2], inplace=True)
                    file.rename(columns={'中国建设银行股份有限公司——复星保德信人寿保险-资本金': '账户1', 'Unnamed: 1': '账户2', 'Unnamed: 2': '币种',
                                         'Unnamed: 3': '昨日余额', 'Unnamed: 4': '借方发生额', 'Unnamed: 5': '借方摘要',
                                         'Unnamed: 6': '借方合计', 'Unnamed: 7': '贷方发生额', 'Unnamed: 8': '贷方摘要',
                                         'Unnamed: 9': '贷方合计', 'Unnamed: 10': '可用余额'}, inplace=True)
                    file = file[file['账户1'].isin(['活期存款账户'])]
                    exec(f'{var_name}=file')
                    exec(f'df_list.append({var_name})')
        return pd.concat(df_list).reset_index(drop=True)


    def 农行_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('_')[2].replace('-',
                                                                             '')  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_excel(path + "/" + i).fillna(0)
                    date = file.columns.to_list()[0].split('_')[0].replace('-', '')
                    file['交易日期'] = date
                    file.rename(columns={file.columns.to_list()[0]: '账户', 'Unnamed: 1': '昨日余额', 'Unnamed: 2': '借方发生额',
                                         'Unnamed: 3': '流入明细', 'Unnamed: 4': '贷方发生额', 'Unnamed: 5': '流出明细',
                                         'Unnamed: 6': '可用余额'}, inplace=True)
                    file = file.loc[(file['昨日余额'] != 0) & (file['昨日余额'] != '前日余额') & (file['昨日余额'] != '0')]
                    exec(f'{var_name}=file')
                    exec(f'df_list.append({var_name})')
        return pd.concat(df_list).reset_index(drop=True)


    def 招商_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('-')[
                        3]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_excel(path + "/" + i, skiprows=4)
                    file.drop(columns=['Unnamed: 5', 'Unnamed: 6', '牛卡号', '股东代码'], inplace=True)
                    file = file[file['资金余额'].notnull()]
                    exec(f'{var_name}=file')
                    exec(f'df_list.append({var_name})')
        return pd.concat(df_list).reset_index(drop=True)


    def 浦发_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = 'Var' + i.split('.')[0].split('_')[3]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_excel(path + "/" + i, skiprows=10, dtype=str)
                    file.drop(columns=['Unnamed: 9'], inplace=True)
                    file = file[file['日期'].notnull()]
                    exec(f'{var_name}=file')
                    exec(f'df_list.append({var_name})')
        return pd.concat(df_list).reset_index(drop=True)


    def QDII_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = "Var" + i.split('.')[0].split('_')[1]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_html(path + "/" + i, flavor='lxml')
                    file[7].drop(file[7].head(1).index, inplace=True)
                    file[7].drop(file[7].tail(1).index, inplace=True)
                    exec(f'{var_name}=file[7]')
                    exec(f'df_list.append({var_name})')
        return pd.concat(df_list).reset_index(drop=True)


    def QDII2_to_df(path, loc_list):
        df_list = []
        for root, dirs, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    var_name = "Var" + i.split('.')[0].split('_')[2]  # 定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator）
                    file = pd.read_html(path + "/" + i, flavor='lxml')
                    file[3].drop(file[3].head(1).index, inplace=True)
                    file[3] = file[3].rename(
                        columns={1: "交易日期", 2: "备注(摘要)", 0: "贷方发生额", 3: "账号", 5: "借方发生额", 6: "可用余额", 4: "账号简称"})
                    file[3]["交易日期"] = file[2].iloc[2, 2].replace('-', '')
                    file[3] = file[3].dropna(subset=["借方发生额", "账号简称"])
                    exec(f'{var_name}=file[3]')
                    exec(f'df_list.append({var_name})')
        return pd.concat(df_list).reset_index(drop=True)


    def alter_extension(old, new, path):
        for root, dirs, files in os.walk(path):
            for i in files:
                if i.split('.')[1] == old:
                    os.rename(path + '/' + i, path + '/' + i.split('.')[0] + '.' + new)


    period = input('输入查询期间：')

    # 工行
    for i in track(['人保', '华泰万能', '华泰资本金', '太保', '工行0752', '申万10号', '申万16号', '资本金外币'], description="工行生成中....."):
        # try:
        #     alter_extension('xls', 'html', Base_route + '/工行/' + i)
        # except:
        #     pass
        exec(f"{i}=工行_to_df(Base_route+'/工行/{i}',['xls','{period}'])")
        exec(f"{i}[[2,3,4]]={i}[[2,3,4]].astype('float')")
        exec(f"{i}[[2,3,4]]={i}[[2,3,4]].astype('str')")
        exec(f"{i}={i}[~{i}[1].isin(['期初余额','期末余额'])]")
        exec(f"{i}={i}" + ".rename(columns={1:'备注(摘要)',2:'借方发生额',3:'贷方发生额',4:'可用余额'}).reset_index(drop=True)")
        print(f"[#838B8B]{i}已完成")

    # 调整为资金系统导入模版格式
    工行_dict = {'工行0752': '1001199429013910752', '华泰万能': '1001199429013911007', '华泰资本金': '1001199429013910504',
               '人保': '1001199429013910876', '申万10号': '1001190729014066622', '申万16号': '1001190729014120673',
               '太保': '1001199429013910628'}
    工行_list = {}
    for i in 工行_dict.items():
        exec(f"{i[0]}['账号']=i[1]")
        exec(f"{i[0]}['账号简称']=i[0]")
        exec(f"{i[0]}={i[0]}[['账号简称','账号','交易日期','借方发生额','贷方发生额','可用余额','备注(摘要)']]")
    print("[bold magenta]COMPLETED [green]:heavy_check_mark:")

    # 建行
    for i in track(range(1), description="建行生成中....."):
        建行 = 建行_to_df(Base_route + '/建行', ['xls', period])
        建行['账号简称'] = '建行'
        建行['账号'] = '31050136360000002074'
        建行.loc[建行['贷方摘要'] != 0, '备注(摘要)'] = 建行.loc[建行['贷方摘要'] != 0, '贷方摘要']
        建行.loc[建行['借方摘要'] != 0, '备注(摘要)'] = 建行.loc[建行['借方摘要'] != 0, '借方摘要']
        建行 = 建行[建行['备注(摘要)'].notnull()]
        建行 = 建行[['账号简称', '账号', '交易日期', '借方发生额', '贷方发生额', '可用余额', '备注(摘要)']]
    print("[bold magenta]COMPLETED [green]:heavy_check_mark:")

    # 农行包括德邦7号，德邦8高，农行7465资本金，农行7473万能，农行7481分红
    for i in track(['德邦7号', '德邦8号', '农行7465资本金', '农行7473万能', '农行7481分红'], description="农行生成中....."):
        if len(period) == 6:
            date = period[0:4] + '-' + period[-2:]
        elif len(period) == 8:
            date = period[0:4] + '-' + period[4:6] + '-' + period[6:8]
        elif len(period) == 4:
            date = period
        exec(f"{i}=农行_to_df(Base_route+'/农行/{i}',['xls','{date}'])")
        print(f"[#838B8B]{i}已完成")

    # 农行
    农行_dict = {'德邦7号': '03340300040027564', '德邦8号': '03340300040027556', '农行7465资本金': '03340300040027465',
               '农行7473万能': '03340300040027473', '农行7481分红': '03340300040027481'}
    农行_list = {}
    for i in 农行_dict.items():
        exec(f"{i[0]}['账号']=i[1]")
        exec(f"{i[0]}['账号简称']=i[0]")
        exec(f"{i[0]}[['流入明细','流出明细']]={i[0]}[['流入明细','流出明细']].astype(str)")
        exec(f"{i[0]}={i[0]}.loc[({i[0]}['借方发生额']!='0') |({i[0]}['贷方发生额']!='0')].reset_index()")
        exec(f"{i[0]}['备注(摘要)']={i[0]}['流入明细']+{i[0]}['流出明细']")
        exec(f"{i[0]}={i[0]}[{i[0]}['账户'].str.contains('上海存款')]")
        exec(f"{i[0]}={i[0]}[['账号简称','账号','交易日期','借方发生额','贷方发生额','可用余额','备注(摘要)','昨日余额']]")
        exec(
            f"{i[0]}={i[0]}.drop('备注(摘要)',axis=1).join({i[0]}['备注(摘要)'].str.split('\\r\\n',expand=True).stack().reset_index(level=1,drop=True).rename('备注(摘要)'))")
        exec(f"{i[0]}={i[0]}[~{i[0]}['备注(摘要)'].isin(['','0'])]")
        try:
            exec(
                f"amount=pd.DataFrame({i[0]}['备注(摘要)'].str.split(':',expand=True).stack().reset_index(level=1,drop=True).rename('备注(摘要)'))")
            amount = amount[~amount['备注(摘要)'].str.contains(u'[\u4e00-\u9fa5]')]
            exec(f"{i[0]}['变动额']=amount['备注(摘要)']")
            exec(
                f"{i[0]}.loc[{i[0]}['备注(摘要)'].str.contains('收'),'借方发生额']={i[0]}.loc[{i[0]}['备注(摘要)'].str.contains('收'),'变动额']")
            exec(f"{i[0]}.loc[{i[0]}['备注(摘要)'].str.contains('收'),'贷方发生额']=0")
            exec(
                f"{i[0]}.loc[{i[0]}['备注(摘要)'].str.contains('付|费|提取'),'贷方发生额']={i[0]}.loc[{i[0]}['备注(摘要)'].str.contains('付|费|提取'),'变动额']")
            exec(f"{i[0]}.loc[{i[0]}['备注(摘要)'].str.contains('付|费|提取'),'借方发生额']=0")
            exec(f"{i[0]}.loc[{i[0]}.duplicated(subset='可用余额',keep='last'),'可用余额']='-'")
        except:
            exec(f"{i[0]}={i[0]}[['账号简称','账号','交易日期','借方发生额','贷方发生额','可用余额','备注(摘要)']]")
        exec(f"{i[0]}={i[0]}[['账号简称','账号','交易日期','借方发生额','贷方发生额','可用余额','备注(摘要)']]")
    print("[bold magenta]COMPLETED [green]:heavy_check_mark:")

    # 招商
    for i in track(range(1), description="招商生成中....."):
        招商 = 招商_to_df(Base_route + '/招商', ['xlsx', period, '普通账户对账单'])
        招商['借方发生额'] = ''
        招商['贷方发生额'] = ''

        # 调整为资金系统导入模版格式
        try:
            招商['贷方发生额'] = 招商.loc[招商['变动金额'].astype(str).str.replace(',', '').astype('float') < 0, '变动金额']
            招商['借方发生额'] = 招商.loc[招商['变动金额'].astype(str).str.replace(',', '').astype('float') > 0, '变动金额']

        except:
            pass

        finally:
            招商['账号简称'] = '招商'
            招商.rename(columns={'摘要': '备注(摘要)', '资金余额': '可用余额', '发生日期': '交易日期'}, inplace=True)

            招商['账号'] = '190001085224'
            for i in ['贷方发生额', '借方发生额']:
                招商[i] = 招商[i].astype(str).str.replace(',', '').astype("float").abs()
            招商 = 招商[['账号简称', '账号', '交易日期', '借方发生额', '贷方发生额', '可用余额', '备注(摘要)']].fillna(0)
            招商['可用余额'] = 招商['可用余额'].str.replace('-', '').fillna(0)
            招商 = 招商[招商['可用余额'] != '0.00']
    print("[bold magenta]COMPLETED [green]:heavy_check_mark:")

    # 浦发
    for i in track(range(1), description="浦发生成中....."):
        浦发 = 浦发_to_df(Base_route + '/浦发', ['xls', period, '每日基金明细表'])
        浦发.drop(columns=["交易时间", "对方户名", "对方账号", "对方开户行"], inplace=True)
        浦发 = 浦发.rename(columns={"日期": "交易日期", "摘要": "备注(摘要)", "汇入": "借方发生额", "汇出": "贷方发生额", "余额": "可用余额"})
        浦发["账号简称"] = "浦发"
        浦发["账号"] = "97990078801880000138"
        order = ["账号简称", "账号", "交易日期", "借方发生额", "贷方发生额", "可用余额", "备注(摘要)"]
        浦发 = 浦发[order]
        浦发 = 浦发[浦发['备注(摘要)'] != " "]
    print("[bold magenta]COMPLETED [green]:heavy_check_mark:")

    for i in track(range(1), description="QDII生成中....."):
        # QDII境内
        # for i in ["QDII"]:
        #     try:
        #         alter_extension('xls', 'html', Base_route + '/工行/' + i)
        #     except:
        #         pass
        file1 = QDII_to_df(Base_route + "/工行/QDII", ['1391', 'xls', period]).copy()
        QDII1373 = file1[file1[0].notnull()]
        QDII1373["账号"] = "1001190729140901391"
        QDII1373["账号简称"] = "QDII1373"
        file2 = QDII_to_df(Base_route + "/工行/QDII", ['3275', 'xls', period]).copy()
        QDII3019 = file2[file2[0].notnull()]
        QDII3019["账号"] = "1001190729130903275"
        QDII3019["账号简称"] = "QDII3019"
        file3 = QDII_to_df(Base_route + "/工行/QDII", ['4503', 'xls', period]).copy()
        QDII4503 = file3[file3[0].notnull()]
        QDII4503["账号"] = "1001202919000124503"
        QDII4503["账号简称"] = "QDII4503"
        QDII4503 = QDII4503.rename(columns={0: "交易日期", 1: "备注(摘要)", 2: "贷方发生额", 3: "借方发生额", 4: "可用余额"})
        QDII1373 = QDII1373.rename(columns={0: "交易日期", 1: "备注(摘要)", 2: "贷方发生额", 3: "借方发生额", 4: "可用余额"})
        QDII3019 = QDII3019.rename(columns={0: "交易日期", 1: "备注(摘要)", 2: "贷方发生额", 3: "借方发生额", 4: "可用余额"})
        order = ["账号简称", "账号", "交易日期", "借方发生额", "贷方发生额", "可用余额", "备注(摘要)"]
        for i in ['QDII4503', 'QDII1373', 'QDII3019']:
            exec(f"{i}[['借方发生额','贷方发生额','可用余额']]={i}[['借方发生额','贷方发生额','可用余额']].astype(float).fillna(0)")

        # QDII境外
        QDII境外 = QDII2_to_df(Base_route + "/工行/QDII", ["Consolidated Cash Report", period])
        for i in range(len(QDII境外.index)):
            if QDII境外.iloc[[i], [5]].values[0][0] == "USD":
                QDII境外.iloc[i, 3] = '861530091922'
                QDII境外.iloc[i, 4] = 'QDII1922'
            else:
                QDII境外.iloc[i, 3] = "861520081104"
                QDII境外.iloc[i, 4] = 'QDII1104'
        order = ["账号简称", "账号", "交易日期", "借方发生额", "贷方发生额", "可用余额", "备注(摘要)"]
        QDII境外[['借方发生额', '贷方发生额', '可用余额']] = QDII境外[['借方发生额', '贷方发生额', '可用余额']].astype(float).fillna(0)
    print("[bold magenta]COMPLETED [green]:heavy_check_mark:")

    for i in ["工行0752", "华泰万能", "华泰资本金", "人保", "申万10号", "申万16号", "太保", "资本金外币", "建行", "德邦7号", "德邦8号", "农行7465资本金", "农行7473万能", "农行7481分红", "招商", "浦发","QDII1373", "QDII3019", "QDII4503", "QDII境外"]:
        try:
            exec(f"{i}['借方发生额']={i}['借方发生额'].str.replace(',','').astype(float)")
        except:
            pass
        try:
            exec(f"{i}['贷方发生额']={i}['贷方发生额'].str.replace(',','').astype(float)")
        except:
            pass
        try:
            exec(f"{i}['可用余额']={i}['可用余额'].str.replace(',','').astype(float)")
        except:
            pass

    全局查询 = pd.concat(
        [工行0752, 华泰万能, 华泰资本金, 人保, 申万10号, 申万16号, 太保, 资本金外币, 建行, 德邦7号, 德邦8号, 农行7465资本金, 农行7473万能, 农行7481分红, 招商, 浦发,QDII1373, QDII3019, QDII4503, QDII境外]).reset_index(drop=True).sort_values(['账号简称', '交易日期']).astype(str)
    工行 = pd.concat(
        [人保, 华泰万能, 华泰资本金, 太保, 工行0752, 申万10号, 申万16号, 资本金外币, QDII1373, QDII3019, QDII4503, QDII境外]).reset_index(drop=True).sort_values(['账号简称', '交易日期']).astype(str)
    QDII = pd.concat([QDII1373, QDII3019, QDII4503, QDII境外]).reset_index(drop=True).sort_values(
        ['账号简称', '交易日期']).astype(str)
    农行 = pd.concat([德邦7号, 德邦8号, 农行7465资本金, 农行7473万能, 农行7481分红]).reset_index(drop=True).sort_values(['账号简称', '交易日期']).astype(str)
    # coding=utf-8

    # 流程图
    tree = Tree("[red]全局查询")
    tree1 = tree.add("[#0000CD]工行")
    tree2 = tree.add("[#32CD32]建行")
    tree3 = tree.add("[#FF00FF]农行")
    tree4 = tree.add("[yellow]招商")
    tree5 = tree.add("[#8B658B]浦发")
    tree1_1 = tree1.add("[#2F4F4F]QDII")
    for i in ['QDII1373', 'QDII3019', 'QDII4503', 'QDII境外']:
        tree1_1.add('[#2F4F4F]' + i)
    for i in ['人保', '华泰万能', '华泰资本金', '太保', '工行0752', '申万10号', '申万16号', '资本金外币']:
        tree1.add('[#2F4F4F]' + i)
    for i in ['德邦7号', '德邦8号', '农行7465资本金', '农行7473万能', '农行7481分红']:
        tree3.add('[#2F4F4F]' + i)

    statement_tree = Padding(tree, (2, 4))
    print(statement_tree)

    while True:
        exec(f"object={input('要查询哪个银行账户？:')}" + ".astype(str)")

        table1 = Table(title="银行头寸查询")
        for i in object.columns:
            if i == '可用余额':
                table1.add_column(i, style='red')
            elif i == '账号简称':
                table1.add_column(i, style='green')
            elif i == '备注(摘要)':
                table1.add_column(i, style='purple')
            elif i == '交易日期':
                table1.add_column(i, style='blue')
            else:
                table1.add_column(i)
        table_centered = Align.center(table1)
        with Live(table_centered, refresh_per_second=1000):
            for i in object.values.tolist():
                time.sleep(0.001)
                table1.add_row(i[0], i[1], i[2], i[3], i[4], i[5], i[6])

        # 导出选项
        Confirm = input('导出到?(system/excel/clipboard/continue/back):')

        if Confirm == 'system':
            try:
                os.remove(Base_route + "/" + period + "资金系统模板.xlsx")
            except:
                pass
            object = object.rename(columns={'借方发生额': "贷方发生额", '贷方发生额': "借方发生额"})
            object.drop(columns=['账号简称'], inplace=True)
            object[['对方银行', "对方户名", "对方账号", "票据类型", "票据号", "对账码", "用途"]] = ''
            资金导出模板 = object[
                ['账号', '交易日期', '借方发生额', '贷方发生额', '可用余额', '对方银行', "对方户名", "对方账号", "票据类型", "票据号", "对账码", '备注(摘要)', "用途"]]
            date = period[0:4] + '-' + period[4:6] + '-' + period[6:8]
            for i in range(len(资金导出模板.index)):
                资金导出模板.iloc[i, 1] = 资金导出模板.iloc[i, 1][0:4] + '-' + 资金导出模板.iloc[i, 1][4:6] + '-' + 资金导出模板.iloc[i, 1][6:8]
            资金导出模板 = 资金导出模板.replace(["0.0", "0.00000000", '0', 0], "")
            资金导出模板.to_excel(Base_route + "/" + period + "资金系统模板.xlsx", index=False)
            console.print("[bold magenta]资金系统模版导出成功 [green]:heavy_check_mark:")

        elif Confirm == 'excel':
            try:
                os.remove(Base_route + '/银行流水.xlsx')
            except:
                pass
            object.to_excel(Base_route + "/" + period + "银行头寸.xlsx", index=False)
            console.print("[bold magenta]导出到Excel成功 [green]:heavy_check_mark:")
            confirm_2 = input('进行下一次查询？：（y/n）')
            if confirm_2 == 'y':
                pass
            else:
                restart_program()

        elif Confirm == 'clipboard':
            object.to_clipboard(index=False)
            console.print("[bold magenta]导出到剪贴板成功 [green]:heavy_check_mark:")
            print("[bold red]alert![/bold red] crtl-v前请提前将excel各列格式改为「文本」")

            confirm_2 = input('进行下一次查询？：（y/n）')
            if confirm_2 == 'y':
                pass
            else:
                restart_program()
        elif Confirm == 'back':
            console.clear()
            restart_program()
        else:
            print(statement_tree)

# ----------------------------------------------------------------------------------------------------------

elif Action_confirm == '交易清算':
    Base_route = '//Volumes//fileshare//投资资产核算管理//01 交易清算//40 每日交易清算'
    console = Console()  # 实例化类
    period = console.input('输入清算期间：')
    if len(period) == 8:
        period = period[0:4] + '-' + period[4:6] + '-' + period[6:8]
    elif len(period) == 6:
        period = period[0:4] + '-' + period[4:6]

    console.clear()

    # 一首小诗
    try:
        Poet = Base_route + '//Poet.md'
        with open(Poet) as MARKDOWN:
            md = Markdown(MARKDOWN.read())
        console.print(md, justify="center")
    except:
        pass


    def files_to_df(path, loc_list):
        import pandas as pd
        import os
        df_list = []
        for root, dir, files in os.walk(path):
            for i in files:
                if all(i.find(f) != -1 for f in loc_list):
                    if path == Base_route + '/德邦':
                        var_name = 'var' + i.split('.')[0].split('[')[1].replace(']', '').replace('-', '')
                    else:
                        var_name = 'var' + i.split('.')[0].replace('-', '').replace('－', '').replace('_',
                                                                                                     '')  # 手动修改—定义动态变量名(str)，不能带运算符（否则会报错：cannot assign to operator)
                    file = pd.read_excel(path + "/" + i)
                    exec(f'{var_name}=file')
                    exec(f'df_list.append({var_name})')
        return pd.concat(df_list)


    # 招商
    def 招商():
        招商 = files_to_df(Base_route + '/招商', ['普通账户对账单', 'xlsx', period])
        招商 = 招商[~招商.iloc[:, 2].isnull()]
        招商.drop(columns=['Unnamed: 1', 'Unnamed: 3', 'Unnamed: 5', 'Unnamed: 6'], inplace=True)
        招商.set_axis(['日期', '市场', '摘要', '数量', '均价', '库存', '金额', '手续费', '印花税', '清算金额', '余额'], axis='columns',
                    inplace=True)
        招商 = 招商[招商['日期'] != '发生日期']
        try:
            招商['交易方式'] = 招商['摘要'].str.split('(', expand=True)[0]
        except:
            招商['交易方式'] = ''
        try:
            招商['证券代码'] = 招商['摘要'].str.split('(', expand=True)[1].str.split(')', expand=True)[0].fillna('-')
        except:
            招商['证券代码'] = ''
        try:
            招商['证券名称'] = 招商['摘要'].str.split('(', expand=True)[1].str.split(')', expand=True)[1].fillna('-')
        except:
            招商['证券名称'] = 招商['摘要']

        for i in ['数量', '均价', '库存', '金额', '手续费', '印花税', '清算金额', '余额']:
            招商[i] = 招商[i].str.replace(',', '').astype('float')
        招商 = 招商.groupby(['日期', '市场', '摘要', '交易方式', '证券代码', '证券名称']).sum().reset_index()
        SZ = 招商[招商['摘要'] != '沪港通组合费'].drop(columns=['摘要', '均价', '余额', '库存']).reset_index(drop=True)
        SZ_fee = 招商[招商['摘要'] == '沪港通组合费'].drop(columns=['摘要', '均价', '余额', '库存']).reset_index(drop=True)
        SZ['组合简称'] = '招商证券'
        SZ = SZ[['日期', '组合简称', '证券代码', '市场', '证券名称', '数量', '手续费', '金额', '印花税', '清算金额', '交易方式']].fillna(0)
        return SZ


    # 申万
    def 申万():
        申万 = files_to_df(Base_route + '/申万', ['成交清算日报表', 'xls', period])
        申万 = 申万[~申万.iloc[:, 0].isnull()]
        申万 = 申万[申万['成交清算日报表'] != '交易日期']
        申万.drop(columns=['Unnamed: 1', 'Unnamed: 5'])
        SW = 申万.iloc[:,
             [0, 2, 3, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 33, 38, 40, 42, 44, 46, 48, 50, 52, 54, 56, 58,
              60, 62, 64, 66, 68, 70, 72]]
        SW = SW.set_axis(
            ['日期', '组合简称', '证券代码0', '证券名称', '数量', '金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '风险金', '其他费用', '债券利息',
             '回购收益', '清算金额', '交易方式', '结算费', '交易费用', '交易手续费', '结算服务费', '上清所交易费', '上清所结算费', '银行手续费', '质押登记费', '认购费',
             '申购费', '赎回费', '交割手续费', '转托管费', '汇款费', '券商过户费', '交易佣金', '交割佣金', '成交价格'], axis='columns')
        SW[['数量', '金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '风险金', '其他费用', '债券利息', '回购收益', '清算金额', '结算费', '交易费用',
            '交易手续费', '结算服务费', '上清所交易费', '上清所结算费', '银行手续费', '质押登记费', '认购费', '申购费', '赎回费', '交割手续费', '转托管费', '汇款费',
            '券商过户费', '交易佣金', '交割佣金', '成交价格']] = SW[
            ['数量', '金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '风险金', '其他费用', '债券利息', '回购收益', '清算金额', '结算费', '交易费用',
             '交易手续费', '结算服务费', '上清所交易费', '上清所结算费', '银行手续费', '质押登记费', '认购费', '申购费', '赎回费', '交割手续费', '转托管费', '汇款费',
             '券商过户费', '交易佣金', '交割佣金', '成交价格']].astype('float')
        for i in ['结算费', '交易费用', '交易手续费', '结算服务费', '上清所交易费', '上清所结算费', '银行手续费', '质押登记费', '认购费', '申购费', '赎回费', '交割手续费',
                  '转托管费', '汇款费', '券商过户费', '交易佣金', '交割佣金']:
            SW['其他费用'] += SW[i]
        SW = SW[['日期', '组合简称', '证券代码0', '证券名称', '数量', '金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '风险金', '其他费用', '债券利息',
                 '回购收益', '清算金额', '交易方式']]
        SW['日期'] = SW['日期'].str.replace('-', '')
        SW['证券代码'] = SW['证券代码0'].str.split(' ', expand=True)[0]
        SW['市场'] = SW['证券代码0'].str.split(' ', expand=True)[1]
        ZQLT = SW[SW['交易方式'] == 'ZQLT']
        SW = SW[SW['交易方式'] != 'ZQLT']
        SW.drop(columns='证券代码0', inplace=True)
        SW = SW[
            ['日期', '组合简称', '证券代码', '市场', '证券名称', '数量', '金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '风险金', '其他费用', '债券利息',
             '回购收益', '清算金额', '交易方式']].fillna(0)
        return SW


    # 德邦
    def 德邦():
        德邦 = files_to_df(Base_route + '/德邦', ['成交清算日报表', 'xls', period])
        德邦 = 德邦.set_axis(
            ['摘要', '证券名称', '市场', '交易方式', '渠道', '数量', '金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '其他费用', '债券利息', '回购收益',
             '清算金额'], axis='columns')
        德邦 = 德邦[德邦.摘要.notnull()]
        德邦 = 德邦[德邦.摘要.str.contains('日期|德邦证券|\d')].reset_index(drop=True)
        德邦['组合简称'] = 德邦.loc[德邦['摘要'].str.contains('德邦证券'), '摘要']
        德邦['日期'] = 德邦.loc[德邦['摘要'].str.contains('日期'), '摘要']
        德邦['证券代码'] = 德邦.loc[德邦['摘要'].str.contains('\d'), '摘要']
        德邦.drop(columns='摘要', inplace=True)
        德邦.组合简称 = 德邦.组合简称.ffill()
        德邦.日期 = 德邦.日期.ffill()
        德邦 = 德邦[德邦['证券代码'].str.contains('\d')]
        德邦 = 德邦[~德邦['证券代码'].str.contains('日期|德邦证券')]
        德邦['日期'] = 德邦['日期'].str.findall('\d').str.join('')
        德邦['组合简称'] = 德邦['组合简称'].str.findall('\d').str.join('')
        德邦.loc[德邦['组合简称'] == '8', '组合简称'] = '德邦8号'
        德邦.loc[德邦['组合简称'] == '7', '组合简称'] = '德邦7号'
        DB = 德邦[
            ['日期', '组合简称', '证券代码', '市场', '证券名称', '数量', '金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '其他费用', '债券利息', '回购收益',
             '清算金额', '交易方式']].fillna(0)
        return DB


    for i in track(range(3), description='生成中....'):
        if i == 0:
            try:
                招商()
                print("[#838B8B]招商已完成")
            except:
                pass
        elif i == 1:
            try:
                申万()
                print("[#838B8B]申万已完成")
            except:
                pass
        elif i == 2:
            try:
                德邦()
            except:
                pass
            print("[#838B8B]德邦已完成")

    try:
        overall = pd.concat([招商(), 申万(), 德邦()]).fillna(0)[
            ['日期', '组合简称', '证券代码', '市场', '交易方式', '证券名称', '数量', '金额', '清算金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '风险金',
             '其他费用', '手续费', '债券利息', '回购收益']]
    except:
        try:
            overall = pd.concat([招商(), 申万()]).fillna(0)[
                ['日期', '组合简称', '证券代码', '市场', '交易方式', '证券名称', '数量', '金额', '清算金额', '席位佣金', '印花税', '经手费', '过户费', '征管费',
                 '风险金', '其他费用', '手续费', '债券利息', '回购收益']]
        except:
            overall = 申万().fillna(0)[
                ['日期', '组合简称', '证券代码', '市场', '交易方式', '证券名称', '数量', '金额', '清算金额', '席位佣金', '印花税', '经手费', '过户费', '征管费',
                 '风险金', '其他费用', '手续费', '债券利息', '回购收益']]
    overall[['数量', '金额', '清算金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '风险金', '其他费用', '手续费', '债券利息', '回购收益']] = overall[
        ['数量', '金额', '清算金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '风险金', '其他费用', '手续费', '债券利息', '回购收益']].astype(float)
    final = overall.groupby(['日期', '组合简称', '证券代码', '市场', '交易方式', '证券名称']).sum().reset_index()

    while True:
        query = period
        date = f"^{query}"

        query_result = final.loc[final['日期'].str.contains(date)]
        describe = query_result.describe().reset_index()
        for i in ['数量', '金额', '清算金额', '席位佣金', '印花税', '经手费', '过户费', '征管费', '风险金', '其他费用', '手续费', '债券利息', '回购收益']:
            query_result[i] = round(query_result[i], 2)
            describe[i] = round(describe[i], 2)

        # 美观输出

        table = Table(show_header=True, header_style="bold magenta")
        for i in query_result.columns:
            if i == '组合简称':
                table.add_column(i, style="cyan")
            elif i == '交易方式':
                table.add_column(i, style="red")
            elif i == '证券名称':
                table.add_column(i, style="yellow")
            else:
                table.add_column(i)
        table_centered = Align.center(table)
        with Live(table_centered, refresh_per_second=1000):
            for i in query_result.astype(str).values.tolist():
                time.sleep(0.001)
                table.add_row(i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8], i[9], i[10], i[11], i[12], i[13],
                              i[14], i[15], i[16], i[17], i[18])

        # 导出选项
        Confirm = input('导出到?(excel/clipboard/back):')
        if Confirm == 'excel':
            query_result.to_excel(Base_route + '/' + query + '交易清算.xlsx', index=False)
            console.print(
                "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
            console.print(
                ' o                 /\' ) \n                      /\'   (                          ,\n                __/\'     )                        .\' `;\n   o      _.-~~~~\'          ``---..__             .\'   ;\n     _.--\'  b)                       ``--...____.\'   .\'\n    (     _.      )).      `-._                     <\n     `\|\|\|\|)-.....___.-     `-.         __...--\'-.\'.\n       `---......____...---`.___.\'----... .\'         `.;\n                                        `-`           `\'')
            confirm_2 = input('进行下一次查询？：（y/n）')
            if confirm_2 == 'y':
                pass
            else:
                restart_program()


        elif Confirm == 'clipboard':
            query_result.to_clipboard(index=False)
            console.print(
                "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
            console.print(
                ' o                 /\' ) \n                      /\'   (                          ,\n                __/\'     )                        .\' `;\n   o      _.-~~~~\'          ``---..__             .\'   ;\n     _.--\'  b)                       ``--...____.\'   .\'\n    (     _.      )).      `-._                     <\n     `\|\|\|\|)-.....___.-     `-.         __...--\'-.\'.\n       `---......____...---`.___.\'----... .\'         `.;\n                                        `-`           `\'')
            if confirm_2 == 'y':
                pass
            else:
                restart_program()

        else:
            restart_program()

# -------------------------------------------------------------------------------------------------------------------------------

elif Action_confirm == '投资款日结报告':
    import openpyxl as op
    import glob
    import pandas as pd
    import os
    import shutil

    Base_route = '//Volumes//fileshare//投资资产核算管理//01 交易清算//50 投资款日结报告'


    def mymovefile(srcfile, dstpath):  # 移动函数
        if not os.path.isfile(srcfile):
            print("%s not exist!" % (srcfile))
        else:
            fpath, fname = os.path.split(srcfile)  # 分离文件名和路径
            if not os.path.exists(dstpath):
                os.makedirs(dstpath)  # 创建路径
            shutil.move(srcfile, dstpath + fname)  # 移动文件
            print("move %s -> %s" % (srcfile, dstpath + fname))


    try:
        os.remove(Base_route + "/策略型产品资产划拨日结报表/策略型产品资产划拨日结报表输入.xlsx")
        print("策略型产品初始化已完成")
    except:
        print("策略型产品无需初始化")
    try:
        os.remove(Base_route + "/分红型产品资产划拨日结报表/分红型产品资产划拨日结报表输入.xlsx")
        print("分红型产品初始化已完成")
    except:
        print("分红型产品无需初始化")
    try:
        os.remove(Base_route + "/投资款日报表/投资款日结报告输入.xlsx")
        print("万能型产品初始化已完成")
    except:
        print("万能型产品无需初始化")

    path = Base_route + "/投资款日报表"
    try:
        tables = []
        for path, dirs, files in os.walk(path):
            for i in files:
                if i.split(".")[1] == "xls":
                    tables.append(i)
        print(tables)  # 读取所有表格名

        os.chdir(Base_route + "/投资款日报表")
        for table_name in tables:

            flag = os.path.isfile(table_name.split(".")[0] + ".csv")
            if not flag:
                os.rename(table_name, "投资款日报表.xls")
                os.system("visidata -p 数据清理.vd --batch")
                os.rename("投资款日报表.xls", table_name)
                exl = op.load_workbook('投资款日报P.xlsx')  # 打开目标Excel文件
                sheet = exl.get_sheet_by_name('投资款日报表_Sheet1_selectedref_T_T')  # 打开该Excel里对应的sheet
                num = 1
                while 1:
                    cell = sheet.cell(row=num, column=2).value
                    if cell:
                        num = num + 1
                    else:
                        break
                for j in range(1, num):  # 对第1至382行单元格遍历（以下是对第二列的操作）
                    if sheet.cell(row=j, column=3).value is not None:  # 如果该单元格不为空
                        value = sheet.cell(row=j, column=3).value  # 那把此格的内容记做value
                    else:  # 如果该单元格为空
                        sheet.cell(j, 3, value)  # 那么填入value
                exl.save(table_name + 'P.xlsx')
                os.remove("投资款日报P.xlsx")
                data_xls = pd.read_excel(table_name + 'P.xlsx', dtype="str")
                data_xls['FUND CODE'] = data_xls['FUND CODE'].astype("str")
                data_xls.to_csv(table_name.split(".")[0] + '.csv')
                os.remove(table_name + 'P.xlsx')
                print(table_name + "已完成")
            else:
                pass

        os.chdir(Base_route + "/投资款日报表")
        try:
            os.remove("combined.csv")
        except:
            pass
        extension = 'csv'
        csv_list = [i for i in glob.glob('*.{}'.format(extension))]
        print(u'共发现%s个CSV文件' % len(csv_list))
        ##合并所有CSV文件
        combined_csv = pd.concat([pd.read_csv(f, dtype=str) for f in csv_list])
        combined_csv['FUND CODE'] = combined_csv['FUND CODE'].astype("str")
        ##导出到CSV
        combined_csv.to_csv("combined.csv", index=False, encoding='utf-8-sig')
        print("合并完成")
        os.system("visidata -p 数据透视.vd --batch")

    except:
        print("万能型产品无数据")

    # 分红型产品资产划拨日结报表
    path = Base_route + "/分红型产品资产划拨日结报表"
    try:
        tables = []
        for path, dirs, files in os.walk(path):
            for i in files:
                if i.split(".")[1] == "xls":
                    tables.append(i)
        print(tables)  # 读取所有表格名

        os.chdir(Base_route + "/分红型产品资产划拨日结报表")
        for table_name in tables:

            flag = os.path.isfile(table_name.split(".")[0] + ".csv")
            if not flag:
                os.rename(table_name, "分红型产品资产划拨日结报表.xls")
                os.system("visidata -p 数据清理.vd --batch")
                os.rename("分红型产品资产划拨日结报表.xls", table_name)
                exl = op.load_workbook('分红型产品资产划拨日结报表P.xlsx')  # 打开目标Excel文件
                sheet = exl.get_sheet_by_name('分红型产品资产划拨日结报表_Sheet1_selectedre')  # 打开该Excel里对应的sheet
                num = 1
                while 1:
                    cell = sheet.cell(row=num, column=2).value
                    if cell:
                        num = num + 1
                    else:
                        break
                for j in range(1, num):  # 对第1至382行单元格遍历（以下是对第二列的操作）
                    if sheet.cell(row=j, column=6).value is not None:  # 如果该单元格不为空
                        value = sheet.cell(row=j, column=6).value  # 那把此格的内容记做value
                    else:  # 如果该单元格为空
                        sheet.cell(j, 6, value)  # 那么填入value
                exl.save(table_name + 'P.xlsx')
                os.remove("分红型产品资产划拨日结报表P.xlsx")
                data_xls = pd.read_excel(table_name + 'P.xlsx', dtype="str")
                data_xls['FUND CODE'] = data_xls['FUND CODE'].astype("str")
                data_xls.to_csv(table_name.split(".")[0] + '.csv')
                os.remove(table_name + 'P.xlsx')
                print(table_name + "已完成")
            else:
                pass

        os.chdir(Base_route + "/分红型产品资产划拨日结报表")
        try:
            os.remove("combined.csv")
        except:
            pass
        extension = 'csv'
        csv_list = [i for i in glob.glob('*.{}'.format(extension))]
        print(u'共发现%s个CSV文件' % len(csv_list))
        ##合并所有CSV文件
        combined_csv = pd.concat([pd.read_csv(f, dtype=str) for f in csv_list])
        combined_csv['FUND CODE'] = combined_csv['FUND CODE'].astype("str")
        ##导出到CSV
        combined_csv.to_csv("combined.csv", index=False, encoding='utf-8-sig')
        print("合并完成")
        os.system("visidata -p 数据透视.vd --batch")

    except:
        print("分红型产品无数据")

    path = Base_route + "/策略型产品资产划拨日结报表"
    try:
        tables = []
        for path, dirs, files in os.walk(path):
            for i in files:
                if i.split(".")[1] == "xls":
                    tables.append(i)
        print(tables)  # 读取所有表格名

        os.chdir(Base_route + "/策略型产品资产划拨日结报表")
        for table_name in tables:

            flag = os.path.isfile(table_name.split(".")[0] + ".csv")
            if not flag:
                os.rename(table_name, "策略型产品资产划拨日结报表.xls")
                os.system("visidata -p 数据清理.vd --batch")
                os.rename("策略型产品资产划拨日结报表.xls", table_name)
                exl = op.load_workbook('策略型产品资产划拨日结报表P.xlsx')  # 打开目标Excel文件
                sheet = exl.get_sheet_by_name('策略型产品资产划拨日结报表_Sheet1_selectedre')  # 打开该Excel里对应的sheet
                num = 1
                while 1:
                    cell = sheet.cell(row=num, column=2).value
                    if cell:
                        num = num + 1
                    else:
                        break
                for j in range(1, num):  # 对第1至382行单元格遍历（以下是对第二列的操作）
                    if sheet.cell(row=j, column=5).value is not None:  # 如果该单元格不为空
                        value = sheet.cell(row=j, column=5).value  # 那把此格的内容记做value
                    else:  # 如果该单元格为空
                        sheet.cell(j, 5, value)  # 那么填入value
                exl.save(table_name + 'P.xlsx')
                os.remove("策略型产品资产划拨日结报表P.xlsx")
                data_xls = pd.read_excel(table_name + 'P.xlsx', dtype="str")
                data_xls['FUND CODE'] = data_xls['FUND CODE'].astype("str")
                data_xls.to_csv(table_name.split(".")[0] + '.csv')
                os.remove(table_name + 'P.xlsx')
                print(table_name + "已完成")
            else:
                pass

        os.chdir(Base_route + "/策略型产品资产划拨日结报表")
        extension = 'csv'
        try:
            os.remove("combined.csv")
        except:
            pass
        csv_list = [i for i in glob.glob('*.{}'.format(extension))]
        print(u'共发现%s个CSV文件' % len(csv_list))
        ##合并所有CSV文件
        combined_csv = pd.concat([pd.read_csv(f, dtype=str) for f in csv_list])
        combined_csv['FUND CODE'] = combined_csv['FUND CODE'].astype("str")
        ##导出到CSV
        combined_csv.to_csv("combined.csv", index=False, encoding='utf-8-sig')
        print("合并完成")

        os.system("visidata -p 数据透视.vd --batch")
    except:
        print("策略型产品无数据")

    os.chdir(Base_route)

    try:
        投资款 = pd.read_excel(Base_route + "/投资款日报表/投资款日结报告输入.xlsx", dtype=str)
        投资款["TYPE"] = "万能"
    except:
        pass
    try:
        分红型 = pd.read_excel(Base_route + "/分红型产品资产划拨日结报表/分红型产品资产划拨日结报表输入.xlsx", dtype=str)
        分红型["TYPE"] = "分红"
    except:
        pass
    try:
        策略型 = pd.read_excel(Base_route + "/策略型产品资产划拨日结报表/策略型产品资产划拨日结报表输入.xlsx", dtype=str)
        策略型["TYPE"] = "策略"
    except:
        pass

    result = pd.concat([投资款, 分红型, 策略型], axis=0, ignore_index=True)

    del result['Unnamed: 0']
    del result['SUBSCRIPTION']
    del result['REDEMPTION']
    result = result.sort_values(by=["TRANSACTION DATE", "FUND CODE"], ascending=True)
    result.loc[result['FUND CODE'] == 'BPEB01', 'FUND CODE'] = "P003"
    result.loc[result['FUND CODE'] == 'BPEB02', 'FUND CODE'] = "P003"
    result.loc[result['FUND CODE'] == 'BPSPEA', 'FUND CODE'] = "P002"
    result.loc[result['FUND CODE'] == 'BPSPEB', 'FUND CODE'] = "P002"
    result.loc[result['FUND CODE'] == 'BNLPA', 'FUND CODE'] = "CL01"
    result.loc[result['FUND CODE'] == 'BNSPA', 'FUND CODE'] = "CL02"
    result = result[~result['SUBSCRIPTION+REDEMPTION'].isin(["0"])]
    result["SUBSCRIPTION+REDEMPTION"] = result["SUBSCRIPTION+REDEMPTION"].astype("float")

    for a in ["投资款日报表", "分红型产品资产划拨日结报表", "策略型产品资产划拨日结报表"]:
        tables_csv = []
        route = Base_route + "/" + a
        os.chdir(Base_route + "/" + a)
        for path, dirs, files in os.walk(route):
            for i in files:
                if i.split(".")[1] == "csv":
                    tables_csv.append(i)
        for files in tables_csv:
            os.remove(files)

    ##统计每日分险种总数
    data_grouped = result.groupby([result["TRANSACTION DATE"], result["TYPE"]]).sum()
    data_grouped = data_grouped.reset_index()

    ##统计每日分FUND CODE总数
    regrouped = result.groupby([result["TRANSACTION DATE"], result['FUND CODE']]).sum()
    regrouped = regrouped.reset_index()

    del result['TYPE']

    # 美观输出
    table = Table(show_header=True, header_style="bold magenta")
    for i in data_grouped.columns:
        table.add_column(i)
    for i in track(data_grouped.astype(str).values.tolist()):
        table.add_row(i[0], i[1], i[2])

    console.print(table)

    # 导出选项
    Confirm = input('导出到?(excel/clipboard/no):')
    if Confirm == 'excel':
        regrouped.to_excel(Base_route + '/TRANSFER WP-活存划拨输入数据.xlsx', index=False)
        console.print(
            "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
        restart_program()
    elif Confirm == 'clipboard':
        result.to_clipboard(index=False)
        console.print(
            "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
        restart_program()
    else:
        console.print(":cry: :cry: :cry: :cry: :cry: :cry: :cry: :cry:")

    restart_program()

# -----------------------------------------------------------------------------------------------------------------------------------------------------------
# 估值表合并
if Action_confirm == '托管行估值表':

    period = input('输入合并月份(yyyymm)')
    period = period[0:4] + '年' + period[4:6] + '月'
    print(period)
    Base_route = '/Volumes/fileshare/投资资产核算管理/03 估值管理/10 月末估值表'


    # def 转化格式(list):

    #     def alter_extension(old, new, path):
    #         global root
    #         for root, dirs, files in os.walk(path):
    #             for i in files:
    #                 if i.split('.')[1] == old:
    #                     os.rename(path + '/' + i, path + '/' + i.split('.')[0] + '.' + new)
    #         print(root + '已转化格式')
    #         for i in list:
    #             alter_extension('xls', 'html', Base_route + '/' + period + '/' + i)


    # 转化格式(['工行'])


    def 工行():

        def 工行处理(file_name):

            df = pd.read_html(file_name, flavor='lxml')
            table = df[5].T.set_index(0).T[['科目代码', '科目名称', '数量', '成本', '市值', '单位成本', '当日均价', '估值增值']]
            var_name = table.loc[(table['成本'] != '0') & (table['科目代码'].str.contains('\d'))][
                ['科目代码', '科目名称', '数量', '成本', '市值']].reset_index(drop=True)
            var_name[['数量', '成本', '市值']] = var_name[['数量', '成本', '市值']].astype(float)
            return var_name

        工行_list = []
        for root, dirs, files in os.walk(Base_route + '/' + period + '/工行'):
            os.chdir(Base_route + '/' + period + '/工行')

            dict = {'复保QDII1104': ['xls', '外币'], '复保工行0752': ['xls', '普通估值表(显示)_复星保德信人寿保险-资本金'],
                    '太保资本金': ['xls', '普通估值表(显示)_复星保德信资本金-太保委托'], '人保资本金': ['xls', '普通估值表(显示)_复星保德信人寿-人保资产'],
                    '华泰万能': ['xls', '普通估值表(显示)_复星保德信人寿-万能险-华泰委托投资户'], '华泰资本金': ['xls', '普通估值表(显示)_复星保德信人寿保险-华泰资产委托']}

            for i in files:
                for x in dict.items():
                    if all(i.find(f) != -1 for f in x[1]):
                        exec(f"{x[0]}=工行处理(i)")
                        exec(f"{x[0]}['投资机构']=x[0]")
                        exec(f"工行_list.append({x[0]})")

        工行 = pd.concat(工行_list)
        return 工行


    def 建行():

        for root, dirs, files in os.walk(Base_route + '/' + period + '/建行'):
            os.chdir(Base_route + '/' + period + '/建行')
            for i in files:
                if all(i.find(f) != -1 for f in ['复星保德信人寿保险-资本金', 'xls']):
                    复保建行2074 = pd.read_excel(i, skiprows=3)
        复保建行2074.rename(columns={'科目格式代码': '科目代码', '科目格式名称': '科目名称'}, inplace=True)
        复保建行2074 = 复保建行2074[['科目代码', '科目名称', '数量', '成本', '市值']]
        复保建行2074 = 复保建行2074.loc[复保建行2074['科目代码'].str.contains('\d')]

        复保建行2074.loc[复保建行2074['科目代码'] == '1301-1110-343-04-TK0388-01', '科目名称'] = '泰康资产-稳盈聚利存款102号'
        复保建行2074.loc[复保建行2074['科目代码'] == '1301-1110-343-04-TK0388-01', '科目代码'] = '301-1110-343-04-TK0388'
        复保建行2074.loc[复保建行2074['科目代码'] == '1301-1110-343-04-WYJL83-01', '科目名称'] = '泰康资产-稳赢聚利83号产品'
        复保建行2074.loc[复保建行2074['科目代码'] == '1301-1110-343-04-WYJL83-01', '科目代码'] = '1301-1110-343-04-WYJL83'
        复保建行2074['投资机构'] = '复保建行2074'

        return 复保建行2074


    def 德邦():
        def 德邦处理(file_name):

            df = pd.read_excel(file_name, skiprows=3)
            df = df.loc[(df['成本'].notnull()) & (df['科目代码'].str.contains('\d'))][
                ['科目代码', '科目名称', '数量', '成本', '市值']].reset_index(drop=True)
            return df

        for root, dirs, files in os.walk(Base_route + '/' + period + '/德邦1,7,8号'):
            os.chdir(Base_route + '/' + period + '/德邦1,7,8号')
            for i in files:
                if all(i.find(f) != -1 for f in ['7号', 'xls']):
                    德邦7号 = 德邦处理(i)
                    德邦7号['投资机构'] = '德邦7号'
                if all(i.find(f) != -1 for f in ['8号', 'xls']):
                    德邦8号 = 德邦处理(i)
                    德邦8号['投资机构'] = '德邦8号'
        德邦 = pd.concat([德邦7号, 德邦8号])

        return 德邦


    def 农行():

        def 农行处理(file_name):
            df = pd.read_excel(file_name, skiprows=3)
            df.rename(columns={'本币成本': '成本', '本币市值': '市值'}, inplace=True)
            df = df.loc[df['成本'].notnull()][['科目代码', '科目名称', '数量', '成本', '市值']].reset_index(drop=True).fillna(0)
            df = df.loc[(df['科目代码'].str.contains('\d')) & (df['科目代码'] != 0)]
            return df

        for root, dirs, files in os.walk(Base_route + '/' + period + '/农行'):
            os.chdir(Base_route + '/' + period + '/农行')
            for i in files:
                if all(i.find(f) != -1 for f in ['万能', 'xls']):
                    农行万能 = 农行处理(i)
                    农行万能['投资机构'] = '复保农行7473'
                if all(i.find(f) != -1 for f in ['资本金', 'xls']):
                    农行资本金 = 农行处理(i)
                    农行资本金['投资机构'] = '复保农行7465'
                if all(i.find(f) != -1 for f in ['xls', '分红']):
                    df = pd.read_excel(i, skiprows=3)
                    df.rename(columns={'本币成本': '成本', '本币市值': '市值'}, inplace=True)
                    农行分红 = df.loc[(df['成本'].notnull()) & (df['科目代码'].str.contains('^\d'))][
                        ['科目代码', '科目名称', '数量', '成本', '市值']].reset_index(drop=True)
                    农行分红['投资机构'] = '复保农行7481'
        农行 = pd.concat([农行万能, 农行资本金, 农行分红])
        农行.loc[农行['数量'].isnull(), '数量'] = 农行.loc[农行['数量'].isnull(), '成本']

        return 农行


    def 浦发():

        def 浦发处理(file_name):
            df = pd.read_excel(file_name, skiprows=3)
            df.rename(columns={'数    量': '数量', '成    本': '成本', '市    值': '市值'}, inplace=True)
            df = df.loc[(df['成本'] != '0') & (df['科目代码'].str.contains('\d'))][
                ['科目代码', '科目名称', '数量', '成本', '市值']].reset_index(drop=True)
            return df

        for root, dirs, files in os.walk(Base_route + '/' + period + '/浦发'):
            os.chdir(Base_route + '/' + period + '/浦发')
            for i in files:
                if all(i.find(f) != -1 for f in ['估值表', 'xlsx', '资本金']):
                    浦发 = 浦发处理(i)
                    浦发['投资机构'] = "复保浦发0138"
        return 浦发


    def 申万():

        def 申万处理(file_name):
            df = pd.read_excel(file_name, skiprows=4)
            df = df.loc[(df['成本'].notnull()) & (df['科目代码'].str.contains('\d')) & (df['成本'].str.contains('\d'))][
                ['科目代码', '科目名称', '数量', '成本', '市值']].reset_index(drop=True)
            return df

        for root, dirs, files in os.walk(Base_route + '/' + period + '/申万10,16号'):
            os.chdir(Base_route + '/' + period + '/申万10,16号')
            for i in files:
                if all(i.find(f) != -1 for f in ['申万', 'xls', '10号']):
                    申万10号 = 申万处理(i)
                    申万10号['投资机构'] = '申万10号'
                if all(i.find(f) != -1 for f in ['申万', 'xls', '16号']):
                    申万16号 = 申万处理(i)
                    申万16号['投资机构'] = '申万16号'
        申万 = pd.concat([申万10号, 申万16号])
        for i in ['数量', '成本', '市值']:
            申万[i] = 申万[i].str.replace(',', '')
            申万[i] = 申万[i].astype('float')
            申万[i] = round(申万[i], 2)

        return 申万


    def 工商亚洲():
        for root, dirs, files in os.walk(Base_route + '/' + period + '/工商亚洲-QDII'):
            os.chdir(Base_route + '/' + period + '/工商亚洲-QDII')
            for i in files:
                if all(i.find(f) != -1 for f in ['xls', 'QDII']):
                    工商亚洲 = pd.read_excel(i, skiprows=5)
                    工商亚洲.rename(columns={'成本值(本币)': '成本', '市值(本币)': '市值'}, inplace=True)
                    工商亚洲['投资机构'] = '复保QDII1104'
                    工商亚洲 = 工商亚洲[['科目代码', '科目名称', '数量', '成本', '市值', '投资机构']]
                    工商亚洲 = 工商亚洲[(工商亚洲['科目代码'].str.contains('\d')) & (工商亚洲['科目名称'].notnull())]
                    for i in ['数量', '成本', '市值']:
                        工商亚洲[i] = 工商亚洲[i].astype('float')
                        工商亚洲[i] = round(工商亚洲[i], 2)

        return 工商亚洲


    工行 = 工行().reset_index(drop=True)
    建行 = 建行().reset_index(drop=True)
    德邦 = 德邦().reset_index(drop=True)
    农行 = 农行().reset_index(drop=True)
    浦发 = 浦发().reset_index(drop=True)
    申万 = 申万().reset_index(drop=True)
    工商亚洲 = 工商亚洲().reset_index(drop=True)
    汇总 = pd.concat([工行, 建行, 德邦, 农行, 浦发, 申万, 工商亚洲]).reset_index(drop=True)

    while True:
        # 美观输出
        exec(f"query={console.input('要查看哪个托管行的估值表？[汇总/工行/复保建行2074/德邦/农行/浦发/申万/工商亚洲]：')}")
        rows_count = query.describe().iloc[0, 0] + 1
        print(rows_count)
        table = Table(title='估值表', show_header=True, header_style="bold magenta")
        for i in query.columns:
            if i == '科目名称':
                table.add_column(i, style='red')
            elif i == '投资机构':
                table.add_column(i, style='cyan')
            else:
                table.add_column(i)
        table_centered = Align.center(table)
        with Live(table_centered, refresh_per_second=1000):
            for i in query.astype(str).values.tolist():
                time.sleep(0.001)
                table.add_row(i[0], i[1], i[2], i[3], i[4], i[5])

        Confirm = input('导出到?(excel/clipboard/no):')
        if Confirm == 'excel':
            query.to_excel(Base_route + '/' + period + '/' + period + '估值表.xlsx', index=False)
            console.print(
                "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
            confirm_2 = input('进行下一次查询？：（y/n）')
            if confirm_2 == 'y':
                pass
            else:
                restart_program()
        elif Confirm == 'clipboard':
            query.to_clipboard(index=False)
            console.print(
                "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
            confirm_2 = input('进行下一次查询？：（y/n）')
            if confirm_2 == 'y':
                pass
            else:
                restart_program()
        else:
            restart_program()


# -----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# 资产持仓查询报表
elif Action_confirm == '资产持仓查询报表':
    Base_route = '/Volumes/fileshare/财务部/20 投资核算处/41 月报输出/01 资产持仓/恒生导出余额表'
    print('[#668B8B]生成中.......')
    hs_db = pd.read_excel(Base_route + "/余额表资产全局查询.xls", dtype=str)
    hs_db[['科目代码', '科目名称']] = hs_db[['科目代码', '科目名称']].fillna(method='ffill', inplace=False)  # 填充[科目代码][科目名称]缺失数据
    hs_db[['本期借方发生', '本期借方数量', '本期借方发生(原币)', '本期贷方发生', '本期贷方数量', '本期贷方发生(原币)', '期末余额', '期末数量', '期末余额(原币)']] = hs_db[
        ['本期借方发生', '本期借方数量', '本期借方发生(原币)', '本期贷方发生', '本期贷方数量', '本期贷方发生(原币)', '期末余额', '期末数量', '期末余额(原币)']].fillna(0)
    hs_overall = hs_db.loc[hs_db['账套'] == '－资产全局查询'].copy()
    hs_details = hs_db.loc[hs_db['账套'] != '－资产全局查询'].copy()

    for i in hs_details.科目级次.unique():
        hs_details.loc[hs_details['科目级次'] == str(i), str(i) + '级科目代码'] = hs_details.loc[
            hs_details['科目级次'] == str(i), '科目代码']
        hs_details.loc[hs_details['科目级次'] == str(i), str(i) + '级科目名称'] = hs_details.loc[
            hs_details['科目级次'] == str(i), '科目名称']
    for i in [1, 2, 3, 4, 5]:
        hs_details.loc[(hs_details[str(i + 1) + '级科目代码'].notnull()) | (hs_details[str(i) + '级科目代码'].notnull())] = \
        hs_details.loc[(hs_details[str(i + 1) + '级科目代码'].notnull()) | (hs_details[str(i) + '级科目代码'].notnull())].fillna(
            method='ffill')
    hs_details.loc[
        hs_details['科目级次'] == '1', ['2级科目代码', '2级科目名称', '3级科目代码', '3级科目名称', '4级科目代码', '4级科目名称', '5级科目代码', '5级科目名称',
                                    '6级科目代码', '6级科目名称']] = ''
    hs_details.loc[hs_details['科目级次'] == '2', ['3级科目代码', '3级科目名称', '4级科目代码', '4级科目名称', '5级科目代码', '5级科目名称', '6级科目代码',
                                               '6级科目名称']] = ''
    hs_details.loc[hs_details['科目级次'] == '3', ['4级科目代码', '4级科目名称', '5级科目代码', '5级科目名称', '6级科目代码', '6级科目名称']] = ''
    hs_details.loc[hs_details['科目级次'] == '4', ['5级科目代码', '5级科目名称', '6级科目代码', '6级科目名称']] = ''
    hs_details.loc[hs_details['科目级次'] == '5', ['6级科目代码', '6级科目名称']] = ''


    # 函数定义
    def get_asset(code):  # 150301-成本|150302-公允价值变动（估值增值）|150303-减值准备|150304-折溢价(利息调整)#| 11320202 应计利息
        name = hs_details.loc[hs_details["科目代码"] == (code), '科目名称'].values.tolist()[0]
        ASS = hs_details.loc[
            (hs_details['科目代码'].str.contains('^' + code)) & (hs_details['科目代码'].str.len() > 13)]  # 资产列表
        ASS = ASS.loc[
            (ASS['科目级次'] == '5') | (ASS['科目级次'] == '6') | (ASS['科目级次'] == '4'), ['账套', '科目级次', '科目代码', '科目名称', '期末数量',
                                                                                 '期末余额', '期末余额(原币)', '币种', '3级科目代码',
                                                                                 '4级科目代码', '5级科目代码', '6级科目代码']]
        ASS = ASS.reset_index(drop=True)
        ASS = ASS.rename(columns={'期末数量': '数量', '期末余额': name, '期末余额(原币)': name + '(原币)'})  # 列名称显示
        ASS[['数量', name, name + '(原币)']] = ASS[['数量', name, name + '(原币)']].astype(float)
        ASS = ASS.groupby(['账套', '科目级次', '科目代码', '科目名称', '币种', '3级科目代码', '4级科目代码', '5级科目代码', '6级科目代码']).sum()
        ASS = ASS.reset_index()

        ASS.loc[ASS['科目级次'] == '6', '资产代码'] = ASS.loc[ASS['科目级次'] == '6', '6级科目代码'].replace(
            ASS['5级科目代码'].unique().tolist(), '', regex=True)
        ASS.loc[ASS['科目级次'] == '5', '资产代码'] = ASS.loc[ASS['科目级次'] == '5', '5级科目代码'].replace(
            ASS['4级科目代码'].unique().tolist(), '', regex=True)
        ASS.loc[ASS['科目级次'] == '4', '资产代码'] = ASS.loc[ASS['科目级次'] == '4', '4级科目代码'].replace(
            ASS['3级科目代码'].unique().tolist(), '', regex=True)
        # 补丁
        if code == '150303':
            ASS.loc[ASS['资产代码'] == '003369', '资产代码'] = '3369'
        # 补丁
        if code not in ['150301', '110101', '150101', '130301', '15110201', '151101', '111101', '211101']:
            del ASS['数量']
            ASS = ASS.groupby(['资产代码', '科目名称', '账套', '币种']).sum().reset_index()
            del ASS['币种']
        else:
            ASS['资产类型描述'] = ''
            ASS.loc[ASS['科目代码'].str.contains('15030101'), '资产类型描述'] = '基金'
            ASS.loc[ASS['科目代码'].str.contains('15030102'), '资产类型描述'] = '股票'
            ASS.loc[ASS['科目代码'].str.contains('15030111'), '资产类型描述'] = '债券'
            ASS.loc[ASS['科目代码'].str.contains('15030112'), '资产类型描述'] = '理财产品'
            ASS.loc[ASS['科目代码'].str.contains('11010133'), '资产类型描述'] = '理财产品'
            ASS.loc[ASS['科目代码'].str.contains('15010101'), '资产类型描述'] = '债券'
            ASS.loc[ASS['科目代码'].str.contains('^130301'), '资产类型描述'] = '理财产品'
            ASS.loc[ASS['科目代码'].str.contains('^151101'), '资产类型描述'] = '投资性房地产'
            ASS.loc[ASS['科目代码'].str.contains('^15110201'), '资产类型描述'] = '股权'
            ASS.loc[ASS['科目代码'].str.contains('^111101'), '资产类型描述'] = '买入返售金融资产'
            # 补丁
            ASS.loc[ASS['资产代码'] == 'BZ20166', '资产类型描述'] = '理财产品'
            # 补丁
            ASS = ASS.groupby(['账套', '资产代码', '科目名称', '资产类型描述', '币种']).sum().reset_index()
        ASS['账套'] = ASS['账套'].str.replace(' ', '')

        return (ASS)


    # 可供出售金融资产
    def FormAFS(code):
        asset = get_asset('150301')
        asset['资产分类'] = '可供出售金融资产'
        for i in code:
            if type(i) == list:
                con_list = []
                for a in i:
                    con_list.append(get_asset(a))
                app = pd.concat(con_list, ignore_index=True).groupby(['账套', '资产代码', '科目名称']).sum()
                asset = pd.merge(asset, app, on=['账套', '资产代码', '科目名称'], how='left')

            else:
                asset = pd.merge(asset, get_asset(i), on=['账套', '资产代码', '科目名称'], how='left')
        asset.drop_duplicates(inplace=True)
        return asset


    AFS = FormAFS(['150302', '150303', '150304', ['11321002', '11320202', '11310202', '11310302'],
                   ['6011', '611101', '611102', '113101']]).fillna(0)
    '''
    AFS['净投资收益']=AFS['利息收入']+AFS['红利收入']  # 计算净投资收益
    AFS['净投资收益(原币)']=AFS['利息收入(原币)']+AFS['红利收入(原币)'] # 计算净投资收益(原币)
    AFS['总投资收益']=AFS['利息收入']+AFS['价差收入']+AFS['红利收入']  # 计算总投资收益（原币）
    AFS['总投资收益(原币)']=AFS['利息收入(原币)']+AFS['价差收入(原币)']+AFS['红利收入(原币)'] # 计算总投资收益(原币)
    AFS['综合投资收益']=AFS['利息收入']+AFS['价差收入']+AFS['红利收入']+AFS['公允价值变动'] # 计算综合投资收益
    '''

    AFS = AFS.rename(columns={'科目名称': '资产描述', '可供出售金融资产': '应计利息', '可供出售金融资产(原币)': '应计利息(原币)', '公允价值变动': '估值增值',
                              '公允价值变动(原币)': '估值增值(原币)', '折溢价': '利息调整', '折溢价(原币)': '利息调整(本币)'})
    AFS['应计利息'] = AFS['应计利息'] + AFS['应收股票股利']
    AFS['应计利息(原币)'] = AFS['应计利息(原币)'] + AFS['应收股票股利(原币)']
    AFS = AFS.loc[AFS['成本'] != 0].drop(
        columns=['利息收入', '利息收入(原币)', '价差收入', '价差收入(原币)', '红利收入', '红利收入(原币)', '应收股票股利', '应收股票股利(原币)'])


    # 交易性金融资产

    def FormTS(code):
        asset = get_asset('110101')
        asset['资产分类'] = '交易性金融资产'
        for i in code:
            if type(i) == list:
                con_list = []
                for a in i:
                    con_list.append(get_asset(a))
                app = pd.concat(con_list, ignore_index=True).groupby(['账套', '资产代码', '科目名称']).sum()
                asset = pd.merge(asset, app, on=['账套', '资产代码', '科目名称'], how='left')

            else:
                asset = pd.merge(asset, get_asset(i), on=['账套', '资产代码', '科目名称'], how='left')
        asset.drop_duplicates(inplace=True)
        return asset


    TS = FormTS(['110103', ['11320201', '11310201'], ['6011', '611101', '611102']])
    TS = pd.merge(TS, get_asset('610101').rename(columns={'交易性金融资产': '公允价值变动损益', '交易性金融资产(原币)': '公允价值变动损益(原币)'}),
                  how='left', on=['资产代码', '科目名称', '账套'])
    TS = TS.rename(columns={'科目名称': '资产描述', '公允价值变动': '估值增值', '公允价值变动(原币)': '估值增值(原币)', '交易性金融资产': '应计利息',
                            '交易性金融资产(原币)': '应计利息(原币)'}).fillna(0)
    TS['资产代码'] = TS['资产代码'].str.replace('0', '')
    '''
    TS['净投资收益']=TS['利息收入']+TS['红利收入']+TS['公允价值变动损益']  # 计算净投资收益
    TS['净投资收益(原币)']=TS['利息收入(原币)']+TS['红利收入(原币)']+TS['公允价值变动损益(原币)'] # 计算净投资收益(原币)
    TS['总投资收益']=TS['利息收入']+TS['价差收入']+TS['红利收入']+TS['公允价值变动损益']   # 计算总投资收益（原币）
    TS['总投资收益(原币)']=TS['利息收入(原币)']+TS['价差收入(原币)']+TS['红利收入(原币)']+TS['公允价值变动损益(原币)']  # 计算总投资收益(原币)
    TS['综合投资收益']=TS['利息收入']+TS['价差收入']+TS['红利收入']+TS['估值增值']+TS['公允价值变动损益']  # 计算综合投资收益
    TS['综合投资收益(原币)']=TS['利息收入(原币)']+TS['价差收入(原币)']+TS['红利收入(原币)']+TS['估值增值(原币)']+TS['公允价值变动损益(原币)']  # 计算综合投资收益(原币)
    '''
    TS = TS.loc[TS['成本'] != 0].drop(
        columns=['利息收入', '利息收入(原币)', '价差收入', '价差收入(原币)', '红利收入', '红利收入(原币)', '公允价值变动损益', '公允价值变动损益(原币)'])


    # 持有至到期投资
    def FormHMS(code):
        asset = get_asset('150101')
        asset['资产分类'] = '持有至到期投资'
        for i in code:
            if type(i) == list:
                con_list = []
                for a in i:
                    con_list.append(get_asset(a))
                app = pd.concat(con_list, ignore_index=True).groupby(['账套', '资产代码', '科目名称']).sum()
                asset = pd.merge(asset, app, on=['账套', '资产代码', '科目名称'], how='left')

            else:
                asset = pd.merge(asset, get_asset(i), on=['账套', '资产代码', '科目名称'], how='left')
        asset.drop_duplicates(inplace=True)
        return asset


    HMS = FormHMS(['150102', '11320203', ['6011', '611101', '611102']])
    '''
    HMS['净投资收益']=HMS['利息收入']+HMS['红利收入']  # 计算净投资收益
    HMS['净投资收益(原币)']=HMS['利息收入(原币)']+HMS['红利收入(原币)'] # 计算净投资收益
    HMS['总投资收益']=HMS['利息收入']+HMS['价差收入']+HMS['红利收入']  # 计算总投资收益
    HMS['总投资收益(原币)']=HMS['利息收入(原币)']+HMS['价差收入(原币)']+HMS['红利收入(原币)'] # 计算总投资收益
    HMS['综合投资收益']=HMS['利息收入']+HMS['价差收入']+HMS['红利收入'] # 计算综合投资收益
    HMS['综合投资收益(原币)']=HMS['利息收入(原币)']+HMS['价差收入(原币)']+HMS['红利收入(原币)'] # 计算综合投资收益(原币)
    '''
    HMS = HMS.rename(columns={'科目名称': '资产描述', '折溢价': '利息调整', '折溢价(原币)': '利息调整(原币)', '持有到期金融资产': '应计利息',
                              '持有到期金融资产(原币)': '应计利息(原币)'}).fillna(0)
    HMS = HMS.loc[HMS['成本'] != 0].drop(columns=['利息收入', '利息收入(原币)', '价差收入', '价差收入(原币)', '红利收入', '红利收入(原币)'])


    # 贷款和应收款项
    def FormLAR(code):
        asset = get_asset('130301')
        # 补丁
        asset.loc[asset['科目名称'] == '华安-义乌综保区基础设施债权投资计划', '科目名称'] = '华安-义乌综保区基础设施债权投资计划（一期）第4期'
        # 补丁 ————华安-义乌综保区基础设施债权投资计划（一期）第4期 在成本和应收利息科目名称不一致问题
        for i in code:
            if type(i) == list:
                con_list = []
                for a in i:
                    con_list.append(get_asset(a))
                app = pd.concat(con_list, ignore_index=True).groupby(['账套', '资产代码', '科目名称']).sum()
                asset = pd.merge(asset, app, on=['账套', '资产代码', '科目名称'], how='left')

            else:
                asset = pd.merge(asset, get_asset(i), on=['账套', '资产代码', '科目名称'], how='left')
        asset.drop_duplicates(inplace=True)
        return asset


    LAR = FormLAR(['1132', ['6011', '611101', '611102']])
    LAR = LAR.rename(columns={'应收利息': '应计利息', '应收利息(原币)': '应计利息(原币)'}).fillna(0)

    # 补丁
    ## 不动产应计利息
    不动产_应计利息 = hs_details.loc[
        (hs_details['科目级次'] == '3') & (hs_details['1级科目代码'] == '1132') & (hs_details['科目代码'].str.len() > 9)].copy()
    不动产_应计利息['账套'] = 不动产_应计利息['账套'].str.replace(' ', '')
    不动产_应计利息['资产代码'] = 不动产_应计利息['3级科目代码']
    不动产_应计利息['资产代码'] = 不动产_应计利息['资产代码'].str.replace('113207', '')
    不动产_应计利息['资产代码'] = 不动产_应计利息['资产代码'].str.replace('113213', '')
    不动产_应计利息 = 不动产_应计利息[['账套', '资产代码', '科目名称', '期末余额', '期末余额(原币)', '币种']]
    不动产_应计利息.rename(columns={'期末余额': '应计利息', '期末余额(原币)': '应计利息(原币)'}, inplace=True)
    不动产_应计利息[['应计利息', '应计利息(原币)']] = 不动产_应计利息[['应计利息', '应计利息(原币)']].astype(float)

    ## 不动产利息收入
    不动产_利息收入 = hs_details.loc[
        (hs_details['科目级次'] == '3') & (hs_details['1级科目代码'] == '6011') & (hs_details['科目代码'].str.len() > 9)].copy()
    不动产_利息收入['资产代码'] = 不动产_利息收入['3级科目代码']
    不动产_利息收入['资产代码'] = 不动产_利息收入['资产代码'].str.replace('601113', '')
    不动产_利息收入['资产代码'] = 不动产_利息收入['资产代码'].str.replace('601108', '')
    不动产_利息收入 = 不动产_利息收入[['账套', '资产代码', '科目名称', '期末余额', '期末余额(原币)', '币种']]
    不动产_利息收入.rename(columns={'期末余额': '利息收入', '期末余额(原币)': '利息收入(原币)'}, inplace=True)
    不动产_利息收入[['利息收入', '利息收入(原币)']] = 不动产_利息收入[['利息收入', '利息收入(原币)']].astype(float)

    LAR = pd.concat([LAR, 不动产_应计利息, 不动产_利息收入]).groupby(['账套', '资产代码', '科目名称', '币种']).sum().reset_index()
    LAR.rename(columns={'科目名称': '资产描述'}, inplace=True)
    LAR = LAR[LAR['数量'] != 0]
    LAR['资产分类'] = '贷款和应收款项'
    LAR['资产类型描述'] = '理财产品'
    # 补丁
    '''
    LAR['净投资收益']=LAR['利息收入']+LAR['红利收入']  # 计算净投资收益
    LAR['净投资收益(原币)']=LAR['利息收入(原币)']+LAR['红利收入(原币)'] # 计算净投资收益
    LAR['总投资收益']=LAR['利息收入']+LAR['价差收入']+LAR['红利收入']  # 计算总投资收益
    LAR['总投资收益(原币)']=LAR['利息收入(原币)']+LAR['价差收入(原币)']+LAR['红利收入(原币)'] # 计算总投资收益
    LAR['综合投资收益']=LAR['利息收入']+LAR['价差收入']+LAR['红利收入'] # 计算综合投资收益
    LAR['综合投资收益(原币)']=LAR['利息收入(原币)']+LAR['价差收入(原币)']+LAR['红利收入(原币)'] # 计算综合投资收益(原币)
    '''
    LAR = LAR.loc[LAR['成本'] != 0].drop(columns=['利息收入', '利息收入(原币)', '价差收入', '价差收入(原币)', '红利收入', '红利收入(原币)'])

    # 长期股权投资
    LTR = pd.concat([get_asset('151101').rename(columns={'成本法': '成本', '成本法(原币)': '成本(原币)'}),
                     get_asset('15110201').rename(columns={'权益法': '成本', '权益法(原币)': '成本(原币)'})])
    LTR = pd.merge(LTR, get_asset('15110202').rename(columns={'损益调整': '总投资收益', '损益调整(原币)': '总投资收益(原币)'}), how='left',
                   on=['资产代码', '科目名称', '账套'])
    '''
    LTR['净投资收益']=0
    LTR['净投资收益(原币)']=0
    LTR['综合投资收益']=LTR['总投资收益']
    LTR['综合投资收益(原币)']=LTR['总投资收益(原币)']
    '''
    LTR['资产分类'] = '长期股权投资'
    LTR.rename(columns={'科目名称': '资产描述'}, inplace=True)
    LTR['资产代码'] = LTR['资产代码'].str.replace('0', '')

    # 其他应付应收款
    ST = hs_details.loc[
        (hs_details['科目代码'] == '1133') | (hs_details['科目代码'] == '3033') | (hs_details['科目代码'] == '1221'), ['账套', '科目名称',
                                                                                                           '期末余额',
                                                                                                           '期末余额(原币)',
                                                                                                           '币种']]
    ST[['期末余额', '期末余额(原币)']] = ST[['期末余额', '期末余额(原币)']].astype(float)
    ST = ST.groupby(['账套', '科目名称', '币种']).sum().reset_index()
    ST.rename(columns={'期末余额': '成本', '期末余额(原币)': '成本(原币)'}, inplace=True)
    ST['资产分类'] = '其他应付应收款'
    ST['资产类型描述'] = '现金'
    ST['科目名称'] = '证券清算款'
    ST['资产代码'] = 'Z60000CK'
    ST.rename(columns={'科目名称': '资产描述'}, inplace=True)
    ST = ST[~ST['成本'].isin([0])]

    # 其他货币资金
    DRB = hs_details.loc[hs_details['科目代码'] == '1021', ['账套', '科目名称', '期末余额', '期末余额(原币)', '币种']]
    DRB[['期末余额', '期末余额(原币)']] = DRB[['期末余额', '期末余额(原币)']].astype(float)
    DRB.rename(columns={'期末余额': '成本', '期末余额(原币)': '成本(原币)'}, inplace=True)
    DRB['资产分类'] = '其他货币资金'
    DRB['资产类型描述'] = '现金'
    DRB['资产代码'] = 'Z40000CK'
    DRB.rename(columns={'科目名称': '资产描述'}, inplace=True)

    # 法定存款

    LR = pd.concat([get_asset('100201').rename(columns={'银行存款': '成本', '银行存款(原币)': '成本(原币)'}),
                    get_asset('154101').rename(columns={'存出资本保证金': '成本', '存出资本保证金(原币)': '成本(原币)'})])
    a = get_asset('113208').rename(columns={'存出保证金': '应计利息', '存出保证金(原币)': '应计利息(原币)'})
    b = get_asset('113201').rename(columns={'银行存款': '应计利息', '银行存款(原币)': '应计利息(原币)'})
    interest = pd.concat([a, b])
    LR = pd.merge(LR, interest, on=['账套', '科目名称'], how='left')
    # LR=pd.merge(LR,get_asset('601101').rename(columns={'银行存款':'净投资收益','银行存款(原币)':'净投资收益(原币)'}),on=['账套','科目名称'],how='left')
    LR = LR[~LR['科目名称'].isin(['工商银行', '工商银行（1007）', '德邦手拉手1号'])].fillna(0)
    LR['币种'] = 'CNY'
    LR['资产类型描述'] = '定期存款'
    LR['资产分类'] = '法定存款'
    del LR['资产代码_x']
    del LR['资产代码_y']
    LR.rename(columns={'科目名称': '资产描述'}, inplace=True)
    LR = LR.loc[LR['成本'] != 0].reset_index(drop=True)
    '''
    LR['总投资收益']=LR['净投资收益']
    LR['总投资收益(原币)']=LR['净投资收益(原币)']
    LR['综合投资收益']=LR['净投资收益']
    LR['综合投资收益(原币)']=LR['净投资收益(原币)']
    '''
    LR['数量'] = LR['成本']
    # 匹配资产代码
    dict = {'1.5亿大额存单浦东分行200221': '银行存款－建行200008283304', '1000万北京洋桥支行190521': '银行存款－建行1003242323',
            '1000万普陀分行200415': '银行存款－建行1003292265', '1000万江苏分行20190927': '银行存款－建行1003262663',
            '1000万济南济大路支行200421': '银行存款－农行1003292528', '1000万浦东分行200414': '银行存款－建行1003292264',
            '1000万浦东分行210924': '银行存款－建行1003384588', '1500万南汇支行200326': '银行存款－中行1003288996'
        , '1亿上海虹桥商务区支行034011': '银行存款－交行1003327274', '2000万上海第二支行210601': '银行存款－建行1003365600',
            '2000万丽园支行210617': '银行存款－民生1003366285', '2000万北京洋桥支行190627': '银行存款－建行1003247283',
            '2000万嘉定支行210608': '银行存款－浦发1003368479', '2000万四川分行210806': '银行存款－交行1003376545',
            '2000万普陀支行181102': '银行存款－中行1003217566', '2000万浦东分行038602': '银行存款－建行1003327276',
            '2000万淮海西路支行914721': '银行存款－中行1003327275',
            '3000万上海自贸区分行191118': '银行存款-上行1003269804', '4000万济南市中支行200420': '银行存款－建行1003292525',
            '5000万北京通州分行190305': '银行存款－建行1003231924', '5000万南通分行190329': '银行存款－中行1003233439',
            '5000万淮海西路支行200420': '银行存款－中行1003292526', '7000万北京石景山支行': '银行存款－建行1003276214',
            '上海分行1亿20200304': '厦门银行200018286731', '兰州分行171031': '浙商银行1003168025',
            '1000万上海虹桥支行180817': '上海浦东发展银行1003205824', '5000万沪西支行180629': '银行存款-中信1003199565',
            '1000万上海花园路支行180813': '银行存款－农行1003205345', '1000万上海分行180612': '宁波银行1003198423',
            '1000万上海滨江支行180926': '银行存款－民生1003211777'}
    for i in dict.items():
        LR.loc[LR['资产描述'] == i[0], '资产代码'] = i[1]

    # 卖出回购/买入返售金融资产
    逆回购 = get_asset('111101').rename(columns={'质押式': '成本', '质押式(原币)': '成本(原币)'})
    逆回购['资产类型描述'] = '逆回购'
    逆回购['资产分类'] = '买入返售金融资产'
    正回购 = get_asset('211101').rename(columns={'质押式': '成本', '质押式(原币)': '成本(原币)'})
    正回购['资产类型描述'] = '正回购'
    正回购['资产分类'] = '卖出回购金融资产'
    RP = pd.concat([正回购, 逆回购])
    # RP=pd.merge(RP,get_asset('6011').rename(columns={'利息收入':'净投资收益','利息收入(原币)':'净投资收益(原币)'}),on=['科目名称','资产代码','账套'],how='left').fillna(0)
    RP.rename(columns={'科目名称': '资产描述'}, inplace=True)
    '''
    RP['总投资收益']=RP['净投资收益']
    RP['总投资收益(原币)']=RP['净投资收益(原币)']
    RP['综合投资收益']=RP['净投资收益']
    RP['综合投资收益(原币)']=RP['净投资收益(原币)']
    '''
    RP = RP.loc[(RP['成本'] != 0) | (RP['资产类型描述'] != '正回购')]
    RP = pd.merge(RP,
                  get_asset('1132').rename(columns={'应收利息': '应计利息', '应收利息(原币)': '应计利息(原币)', '科目名称': '资产描述'}).fillna(0),
                  how='left', on=['资产代码', '资产描述', '账套'])

    # 数据汇总
    资产持仓 = pd.concat([AFS, TS, HMS, LAR, LTR, ST, DRB, LR, RP]).fillna(0)
    资产持仓['账套'] = 资产持仓['账套'].str.replace(' ', '')
    资产持仓.loc[资产持仓['资产代码'] == 0, '资产代码'] = ''

    Type = input('分账户 or 分账套？：')
    if Type == '分账户':
        资产持仓.loc[资产持仓['账套'].str.contains('200018_财富鸿盈A|200019_财富鸿盈B|200024_财富享盈A|200025_财富享盈B'), '账套'] = '分红'
        资产持仓.loc[资产持仓['账套'].str.contains(
            '3003_天天盈_太保委托|2007_天天盈A-自营|3006_天天盈B-太保委托|3008_天天盈B-活动|8001_万能归集账户|200005_稳健型子账户|200004_附加天天盈B（钻石）|200009_附加天天盈B（铂金）|200008_万能归集账户2|200013_天添利年金|200012_天天盈B（紫金）|200016_天天盈年金|200017_天添利B|200022_天天盈年金（尊享版）|200026_天天盈年金（臻享版）|200027_天添利B款（尊耀）'), '账套'] = '万能'
        资产持仓.loc[资产持仓['账套'].str.contains(
            '1003_资本金_自营|2001_资本金-华泰委托|3007_资本金-太保委托|4003_资本金-人保委托|5004_资本金-QDII境外投资|200001_德邦-手拉手7号|200002_德邦-手拉手8号|200010_申万宏源共盈10号|200014_自营-建行托管|200020_申万宏源共盈16号|200021_浦发-自营|200007_财富固盈年金|200015_财富稳赢|200011_海通资管-01'), '账套'] = '资本金'
        资产持仓 = 资产持仓.groupby(['资产代码', '资产描述', '账套', '资产类型描述', '资产分类', '币种']).sum().sort_values(
            ['资产分类', '资产描述', '账套', '资产类型描述', ]).reset_index()
    else:
        pass

    # 计算市值
    资产持仓['市值'] = 资产持仓['成本'] + 资产持仓['估值增值'] + 资产持仓['利息调整'] + 资产持仓['减值准备']
    资产持仓.loc[资产持仓['资产分类'] == '长期股权投资', '市值'] += 资产持仓.loc[资产持仓['资产分类'] == '长期股权投资', '总投资收益']

    资产持仓['价格'] = round(资产持仓['市值'] / 资产持仓['数量'], 2).replace(np.inf, 0)  # 计算价格

    资产持仓 = 资产持仓.loc[
        资产持仓['成本'] != 0, ['账套', '资产代码', '资产描述', '资产类型描述', '资产分类', '币种', '价格', '数量', '成本', '市值', '估值增值', '减值准备', '利息调整',
                          '应计利息']]

    # 数据导出

    资产持仓.loc[资产持仓['资产描述'] == '宁波勤邦新材料科技有限公司', '资产代码'] = 'NBQB'
    资产持仓.loc[资产持仓['资产描述'] == '中再锐驰2号资产管理产品', '资产代码'] = 'zzzc-01'
    try:
        adjust = pd.read_excel(Base_route + '/固定调整项目.xlsx')
        资产持仓 = pd.concat([adjust, 资产持仓]).fillna(0)
    except:
        pass

    while True:
        # 美观输出

        table = Table(title='资产持仓查询报表', show_header=True, header_style="bold magenta")
        for i in 资产持仓.columns:
            if i == '资产描述':
                table.add_column(i, style='red')
            elif i == '账套':
                table.add_column(i, style='cyan')
            else:
                table.add_column(i)
        table_centered = Align.center(table)
        with Live(table_centered, refresh_per_second=1000):
            for i in 资产持仓.astype(str).values.tolist():
                time.sleep(0.001)
                table.add_row(i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8], i[9], i[10], i[11], i[12], i[13])

        Confirm = input('导出到?(excel/clipboard/no):')
        if Confirm == 'excel':
            资产持仓.to_excel(Base_route + '/资产持仓查询报表.xlsx', index=False)
            console.print(
                "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
            restart_program()

        elif Confirm == 'clipboard':
            资产持仓.to_clipboard(index=False)
            console.print(
                "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
        else:
            restart_program()

# ——————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————
elif Action_confirm == '资产收益查询报表':
    Base_route = '/Volumes/fileshare/0跨部门共享/投资资产核算管理/06 财务输出报表/资产持仓数据源'

    hs_db = pd.read_excel(Base_route + "/余额表资产全局查询.xls", dtype=str)
    hs_db[['科目代码', '科目名称']] = hs_db[['科目代码', '科目名称']].fillna(method='ffill', inplace=False)  # 填充[科目代码][科目名称]缺失数据
    hs_db[['本期借方发生', '本期借方数量', '本期借方发生(原币)', '本期贷方发生', '本期贷方数量', '本期贷方发生(原币)', '期末余额', '期末数量', '期末余额(原币)']] = hs_db[
        ['本期借方发生', '本期借方数量', '本期借方发生(原币)', '本期贷方发生', '本期贷方数量', '本期贷方发生(原币)', '期末余额', '期末数量', '期末余额(原币)']].fillna(0)
    hs_overall = hs_db.loc[hs_db['账套'] == '－资产全局查询'].copy()
    hs_details = hs_db.loc[hs_db['账套'] != '－资产全局查询'].copy()

    for i in hs_details.科目级次.unique():
        hs_details.loc[hs_details['科目级次'] == str(i), str(i) + '级科目代码'] = hs_details.loc[
            hs_details['科目级次'] == str(i), '科目代码']
        hs_details.loc[hs_details['科目级次'] == str(i), str(i) + '级科目名称'] = hs_details.loc[
            hs_details['科目级次'] == str(i), '科目名称']
    for i in [1, 2, 3, 4, 5]:
        hs_details.loc[(hs_details[str(i + 1) + '级科目代码'].notnull()) | (hs_details[str(i) + '级科目代码'].notnull())] = \
        hs_details.loc[(hs_details[str(i + 1) + '级科目代码'].notnull()) | (hs_details[str(i) + '级科目代码'].notnull())].fillna(
            method='ffill')
    hs_details.loc[
        hs_details['科目级次'] == '1', ['2级科目代码', '2级科目名称', '3级科目代码', '3级科目名称', '4级科目代码', '4级科目名称', '5级科目代码', '5级科目名称',
                                    '6级科目代码', '6级科目名称']] = ''
    hs_details.loc[hs_details['科目级次'] == '2', ['3级科目代码', '3级科目名称', '4级科目代码', '4级科目名称', '5级科目代码', '5级科目名称', '6级科目代码',
                                               '6级科目名称']] = ''
    hs_details.loc[hs_details['科目级次'] == '3', ['4级科目代码', '4级科目名称', '5级科目代码', '5级科目名称', '6级科目代码', '6级科目名称']] = ''
    hs_details.loc[hs_details['科目级次'] == '4', ['5级科目代码', '5级科目名称', '6级科目代码', '6级科目名称']] = ''
    hs_details.loc[hs_details['科目级次'] == '5', ['6级科目代码', '6级科目名称']] = ''


    def get_asset(code):  # 150301-成本|150302-公允价值变动（估值增值）|150303-减值准备|150304-折溢价(利息调整)#| 11320202 应计利息
        name = hs_details.loc[hs_details["科目代码"] == (code), '科目名称'].values.tolist()[0]
        ASS = hs_details.loc[
            (hs_details['科目代码'].str.contains('^' + code)) & (hs_details['科目代码'].str.len() > 13)]  # 资产列表
        ASS = ASS.loc[
            (ASS['科目级次'] == '5') | (ASS['科目级次'] == '6') | (ASS['科目级次'] == '4'), ['账套', '科目级次', '科目代码', '科目名称', '期末数量',
                                                                                 '期末余额', '期末余额(原币)', '币种', '3级科目代码',
                                                                                 '4级科目代码', '5级科目代码', '6级科目代码']]
        ASS = ASS.reset_index(drop=True)
        ASS = ASS.rename(columns={'期末数量': '数量', '期末余额': name, '期末余额(原币)': name + '(原币)'})  # 列名称显示
        ASS[['数量', name, name + '(原币)']] = ASS[['数量', name, name + '(原币)']].astype(float)
        ASS = ASS.groupby(['账套', '科目级次', '科目代码', '科目名称', '币种', '3级科目代码', '4级科目代码', '5级科目代码', '6级科目代码']).sum()
        ASS = ASS.reset_index()

        ASS.loc[ASS['科目级次'] == '6', '资产代码'] = ASS.loc[ASS['科目级次'] == '6', '6级科目代码'].replace(
            ASS['5级科目代码'].unique().tolist(), '', regex=True)
        ASS.loc[ASS['科目级次'] == '5', '资产代码'] = ASS.loc[ASS['科目级次'] == '5', '5级科目代码'].replace(
            ASS['4级科目代码'].unique().tolist(), '', regex=True)
        ASS.loc[ASS['科目级次'] == '4', '资产代码'] = ASS.loc[ASS['科目级次'] == '4', '4级科目代码'].replace(
            ASS['3级科目代码'].unique().tolist(), '', regex=True)
        # 补丁
        if code == '150303':
            ASS.loc[ASS['资产代码'] == '003369', '资产代码'] = '3369'
        # 补丁
        if code not in ['150301', '110101', '150101', '130301', '15110201', '151101', '111101', '211101']:
            del ASS['数量']
            ASS = ASS.groupby(['资产代码', '科目名称', '账套', '币种']).sum().reset_index()
            del ASS['币种']
        else:
            ASS['资产类型描述'] = ''
            ASS.loc[ASS['科目代码'].str.contains('15030101'), '资产类型描述'] = '基金'
            ASS.loc[ASS['科目代码'].str.contains('15030102'), '资产类型描述'] = '股票'
            ASS.loc[ASS['科目代码'].str.contains('15030111'), '资产类型描述'] = '债券'
            ASS.loc[ASS['科目代码'].str.contains('15030112'), '资产类型描述'] = '理财产品'
            ASS.loc[ASS['科目代码'].str.contains('11010133'), '资产类型描述'] = '理财产品'
            ASS.loc[ASS['科目代码'].str.contains('15010101'), '资产类型描述'] = '债券'
            ASS.loc[ASS['科目代码'].str.contains('^130301'), '资产类型描述'] = '理财产品'
            ASS.loc[ASS['科目代码'].str.contains('^151101'), '资产类型描述'] = '投资性房地产'
            ASS.loc[ASS['科目代码'].str.contains('^15110201'), '资产类型描述'] = '股权'
            ASS.loc[ASS['科目代码'].str.contains('^111101'), '资产类型描述'] = '买入返售金融资产'
            # 补丁
            ASS.loc[ASS['资产代码'] == 'BZ20166', '资产类型描述'] = '理财产品'
            # 补丁
            ASS = ASS.groupby(['账套', '资产代码', '科目名称', '资产类型描述', '币种']).sum().reset_index()
        ASS['账套'] = ASS['账套'].str.replace(' ', '')

        return (ASS)


    # 可供出售金融资产
    def FormAFS(code):
        asset = get_asset('150301')
        asset['资产分类'] = '可供出售金融资产'
        for i in code:
            if type(i) == list:
                con_list = []
                for a in i:
                    con_list.append(get_asset(a))
                app = pd.concat(con_list, ignore_index=True).groupby(['账套', '资产代码', '科目名称']).sum()
                asset = pd.merge(asset, app, on=['账套', '资产代码', '科目名称'], how='left')

            else:
                asset = pd.merge(asset, get_asset(i), on=['账套', '资产代码', '科目名称'], how='left')
        asset.drop_duplicates(inplace=True)
        return asset


    AFS = FormAFS(
        [['60110602', '60111002', '61110203', '61110204', '61110206'], ['61110105', '61110106', '61110107', '61110113'],
         ['4002']]).fillna(0)
    净投资收益_AFS = AFS.iloc[:, 9] + AFS.iloc[:, 11] + AFS.iloc[:, 13] + AFS.iloc[:, 15] + AFS.iloc[:, 17]
    总投资收益_AFS = AFS.iloc[:, 9] + AFS.iloc[:, 11] + AFS.iloc[:, 13] + AFS.iloc[:, 15] + AFS.iloc[:, 17] + AFS.iloc[:,
                                                                                                         19] + AFS.iloc[
                                                                                                               :,
                                                                                                               21] + AFS.iloc[
                                                                                                                     :,
                                                                                                                     23] + AFS.iloc[
                                                                                                                           :,
                                                                                                                           25]
    综合投资收益_AFS = AFS.iloc[:, 9] + AFS.iloc[:, 11] + AFS.iloc[:, 13] + AFS.iloc[:, 15] + AFS.iloc[:, 17] + AFS.iloc[:,
                                                                                                          19] + AFS.iloc[
                                                                                                                :,
                                                                                                                21] + AFS.iloc[
                                                                                                                      :,
                                                                                                                      23] + AFS.iloc[
                                                                                                                            :,
                                                                                                                            25] + AFS.iloc[
                                                                                                                                  :,
                                                                                                                                  27]
    AFS['净投资收益'] = 净投资收益_AFS
    AFS['总投资收益'] = 总投资收益_AFS
    AFS['综合投资收益'] = 综合投资收益_AFS
    AFS = AFS[['账套', '资产代码', '科目名称', '资产类型描述', '币种', '数量', '成本', '资产分类', '净投资收益', '总投资收益', '综合投资收益']]
    AFS = AFS.loc[(AFS['净投资收益'] != 0) | (AFS['总投资收益'] != 0) | (AFS['综合投资收益'] != 0)]


    # 交易性金融资产
    def FormTS(code):
        asset = get_asset('110101')
        asset['资产分类'] = '交易性金融资产'
        for i in code:
            if type(i) == list:
                con_list = []
                for a in i:
                    con_list.append(get_asset(a))
                app = pd.concat(con_list, ignore_index=True).groupby(['账套', '资产代码', '科目名称']).sum()
                asset = pd.merge(asset, app, on=['账套', '资产代码', '科目名称'], how='left')

            else:
                asset = pd.merge(asset, get_asset(i), on=['账套', '资产代码', '科目名称'], how='left')
        asset.drop_duplicates(inplace=True)
        return asset


    TS = FormTS([['61110201', '61110202'], ['61110102', '61110103'], ['6101']])
    TS = TS.fillna(0)
    净投资收益_TS = TS.iloc[:, 9] + TS.iloc[:, 11]
    总投资收益_TS = TS.iloc[:, 9] + TS.iloc[:, 11] + TS.iloc[:, 13] + TS.iloc[:, 15]
    综合投资收益_TS = TS.iloc[:, 9] + TS.iloc[:, 11] + TS.iloc[:, 13] + TS.iloc[:, 15] + TS.iloc[:, 17]
    TS['净投资收益'] = 净投资收益_TS
    TS['总投资收益'] = 总投资收益_TS
    TS['综合投资收益'] = 综合投资收益_TS
    TS = TS[['账套', '资产代码', '科目名称', '资产类型描述', '币种', '数量', '成本', '资产分类', '净投资收益', '总投资收益', '综合投资收益']].fillna(0)
    TS = TS.loc[(TS['净投资收益'] != 0) | (TS['总投资收益'] != 0) | (TS['综合投资收益'] != 0)]


    # 持有至到期投资
    def FormHMS(code):
        asset = get_asset('150101')
        asset['资产分类'] = '持有至到期投资'
        for i in code:
            if type(i) == list:
                con_list = []
                for a in i:
                    con_list.append(get_asset(a))
                app = pd.concat(con_list, ignore_index=True).groupby(['账套', '资产代码', '科目名称']).sum()
                asset = pd.merge(asset, app, on=['账套', '资产代码', '科目名称'], how='left')

            else:
                asset = pd.merge(asset, get_asset(i), on=['账套', '资产代码', '科目名称'], how='left')
        asset.drop_duplicates(inplace=True)
        return asset


    HMS = FormHMS(['60110603', '61110109'])
    HMS = HMS.fillna(0)
    净投资收益_HMS = HMS.iloc[:, 9]
    总投资收益_HMS = HMS.iloc[:, 9] + HMS.iloc[:, 11]
    综合投资收益_HMS = HMS.iloc[:, 9] + HMS.iloc[:, 11]
    HMS['净投资收益'] = 净投资收益_HMS
    HMS['总投资收益'] = 总投资收益_HMS
    HMS['综合投资收益'] = 综合投资收益_HMS
    HMS = HMS[['账套', '资产代码', '科目名称', '资产类型描述', '币种', '数量', '成本', '资产分类', '净投资收益', '总投资收益', '综合投资收益']].fillna(0)
    HMS = HMS.loc[(HMS['净投资收益'] != 0) | (HMS['总投资收益'] != 0) | (HMS['综合投资收益'] != 0)]


    # 贷款和应收款项
    def FormLAR(code):
        asset = get_asset('130301')
        asset['资产分类'] = '贷款和应收款项'
        # 补丁
        asset.loc[asset['科目名称'] == '华安-义乌综保区基础设施债权投资计划', '科目名称'] = '华安-义乌综保区基础设施债权投资计划（一期）第4期'
        # 补丁 ————华安-义乌综保区基础设施债权投资计划（一期）第4期 在成本和应收利息科目名称不一致问题
        for i in code:
            if type(i) == list:
                con_list = []
                for a in i:
                    con_list.append(get_asset(a))
                app = pd.concat(con_list, ignore_index=True).groupby(['账套', '资产代码', '科目名称']).sum()
                asset = pd.merge(asset, app, on=['账套', '资产代码', '科目名称'], how='left')

            else:
                asset = pd.merge(asset, get_asset(i), on=['账套', '资产代码', '科目名称'], how='left')
        asset.drop_duplicates(inplace=True)
        return asset


    LAR = FormLAR(['601107', '60111004', '601111', '601112', '601113']).fillna(0)
    净投资收益_LAR = LAR.iloc[:, 9] + LAR.iloc[:, 11] + LAR.iloc[:, 13] + LAR.iloc[:, 15] + LAR.iloc[:, 17]
    总投资收益_LAR = LAR.iloc[:, 9] + LAR.iloc[:, 11] + LAR.iloc[:, 13] + LAR.iloc[:, 15] + LAR.iloc[:, 17]
    综合投资收益_LAR = LAR.iloc[:, 9] + LAR.iloc[:, 11] + LAR.iloc[:, 13] + LAR.iloc[:, 15] + LAR.iloc[:, 17]
    LAR['净投资收益'] = 净投资收益_LAR
    LAR['总投资收益'] = 总投资收益_LAR
    LAR['综合投资收益'] = 综合投资收益_LAR
    LAR = LAR[['账套', '资产代码', '科目名称', '资产类型描述', '币种', '数量', '成本', '资产分类', '净投资收益', '总投资收益', '综合投资收益']].fillna(0)
    LAR = LAR.loc[(LAR['净投资收益'] != 0) | (LAR['总投资收益'] != 0) | (LAR['综合投资收益'] != 0)]

    不动产_利息收入 = hs_details.loc[
        (hs_details['科目级次'] == '3') & (hs_details['1级科目代码'] == '6011') & (hs_details['科目代码'].str.len() > 9)].copy()
    不动产_利息收入['资产代码'] = 不动产_利息收入['3级科目代码']
    不动产_利息收入 = 不动产_利息收入[不动产_利息收入['资产代码'].str.contains('601107|60111004|601111|601112|601113')]
    for i in ['601107', '60111004', '601111', '601112', '601113']:
        不动产_利息收入['资产代码'] = 不动产_利息收入['资产代码'].str.replace(i, '')

    不动产_利息收入 = 不动产_利息收入[['账套', '资产代码', '科目名称', '期末余额', '期末余额(原币)', '币种']]
    不动产_利息收入.rename(columns={'期末余额': '利息收入', '期末余额(原币)': '利息收入(原币)'}, inplace=True)
    不动产_利息收入[['利息收入', '利息收入(原币)']] = 不动产_利息收入[['利息收入', '利息收入(原币)']].astype(float)
    不动产_利息收入.rename(columns={'利息收入': '净投资收益'}, inplace=True)
    不动产_利息收入['总投资收益'] = 不动产_利息收入['净投资收益']
    不动产_利息收入['综合投资收益'] = 不动产_利息收入['净投资收益']
    不动产_利息收入 = 不动产_利息收入[['账套', '资产代码', '科目名称', '币种', '净投资收益', '总投资收益', '综合投资收益']].fillna(0)
    不动产_利息收入 = 不动产_利息收入.loc[(不动产_利息收入['净投资收益'] != 0) | (不动产_利息收入['总投资收益'] != 0) | (不动产_利息收入['综合投资收益'] != 0)]
    不动产_利息收入['资产分类'] = '贷款和应收款项'
    不动产_利息收入['资产类型描述'] = '理财产品'
    LAR = pd.concat([LAR, 不动产_利息收入]).groupby(['账套', '资产代码', '科目名称', '币种', '资产分类', '资产类型描述']).sum().reset_index()

    # 长期股权投资
    LTR = pd.concat([get_asset('151101').rename(columns={'成本法': '成本', '成本法(原币)': '成本(原币)'}),
                     get_asset('15110201').rename(columns={'权益法': '成本', '权益法(原币)': '成本(原币)'})])
    LTR = pd.merge(LTR, get_asset('15110202').rename(columns={'损益调整': '总投资收益', '损益调整(原币)': '总投资收益(原币)'}), how='left',
                   on=['资产代码', '科目名称', '账套'])
    LTR['净投资收益'] = 0
    LTR['综合投资收益'] = LTR['总投资收益']
    LTR['资产分类'] = '长期股权投资'
    LTR = LTR[['账套', '资产代码', '科目名称', '资产类型描述', '币种', '数量', '成本', '资产分类', '净投资收益', '总投资收益', '综合投资收益']].fillna(0)

    # 法定存款
    LR = pd.concat([get_asset('100201').rename(columns={'银行存款': '成本', '银行存款(原币)': '成本(原币)'}),
                    get_asset('154101').rename(columns={'存出资本保证金': '成本', '存出资本保证金(原币)': '成本(原币)'})])
    for i in ['60110102', '60110103', '60110104', '60110107']:
        LR = pd.merge(LR, get_asset(i), on=['账套', '科目名称'], how='left')
    LR = LR[~LR['科目名称'].isin(['工商银行', '工商银行（1007）', '德邦手拉手1号'])].fillna(0)
    LR['币种'] = 'CNY'
    LR['资产类型描述'] = '定期存款'
    LR['资产分类'] = '法定存款'
    del LR['资产代码_x']
    del LR['资产代码_y']
    LR['数量'] = LR['成本']
    净投资收益_LR = LR.iloc[:, 4] + LR.iloc[:, 6] + LR.iloc[:, 8] + LR.iloc[:, 11]
    LR['净投资收益'] = 净投资收益_LR
    LR['总投资收益'] = 净投资收益_LR
    LR['综合投资收益'] = 净投资收益_LR
    dict = {'1.5亿大额存单浦东分行200221': '银行存款－建行200008283304', '1000万北京洋桥支行190521': '银行存款－建行1003242323',
            '1000万普陀分行200415': '银行存款－建行1003292265', '1000万江苏分行20190927': '银行存款－建行1003262663',
            '1000万济南济大路支行200421': '银行存款－农行1003292528', '1000万浦东分行200414': '银行存款－建行1003292264',
            '1000万浦东分行210924': '银行存款－建行1003384588', '1500万南汇支行200326': '银行存款－中行1003288996'
        , '1亿上海虹桥商务区支行034011': '银行存款－交行1003327274', '2000万上海第二支行210601': '银行存款－建行1003365600',
            '2000万丽园支行210617': '银行存款－民生1003366285', '2000万北京洋桥支行190627': '银行存款－建行1003247283',
            '2000万嘉定支行210608': '银行存款－浦发1003368479', '2000万四川分行210806': '银行存款－交行1003376545',
            '2000万普陀支行181102': '银行存款－中行1003217566', '2000万浦东分行038602': '银行存款－建行1003327276',
            '2000万淮海西路支行914721': '银行存款－中行1003327275',
            '3000万上海自贸区分行191118': '银行存款-上行1003269804', '4000万济南市中支行200420': '银行存款－建行1003292525',
            '5000万北京通州分行190305': '银行存款－建行1003231924', '5000万南通分行190329': '银行存款－中行1003233439',
            '5000万淮海西路支行200420': '银行存款－中行1003292526', '7000万北京石景山支行': '银行存款－建行1003276214',
            '上海分行1亿20200304': '厦门银行200018286731', '兰州分行171031': '浙商银行1003168025',
            '1000万上海虹桥支行180817': '上海浦东发展银行1003205824', '5000万沪西支行180629': '银行存款-中信1003199565',
            '1000万上海花园路支行180813': '银行存款－农行1003205345', '1000万上海分行180612': '宁波银行1003198423',
            '1000万上海滨江支行180926': '银行存款－民生1003211777', '兰州分行211029': '浙商银行1003390189'}
    for i in dict.items():
        LR.loc[LR['科目名称'] == i[0], '资产代码'] = i[1]
    LR = LR[['账套', '资产代码', '科目名称', '资产类型描述', '币种', '数量', '成本', '资产分类', '净投资收益', '总投资收益', '综合投资收益']].fillna(0)
    LR = LR.loc[(LR['净投资收益'] != 0) | (LR['总投资收益'] != 0) | (LR['综合投资收益'] != 0)]

    # 卖出回购/买入返售金融资产
    逆回购 = get_asset('111101').rename(columns={'质押式': '成本', '质押式(原币)': '成本(原币)'})
    逆回购['资产类型描述'] = '逆回购'
    逆回购['资产分类'] = '买入返售金融资产'
    正回购 = get_asset('211101').rename(columns={'质押式': '成本', '质押式(原币)': '成本(原币)'})
    正回购['资产类型描述'] = '正回购'
    正回购['资产分类'] = '卖出回购金融资产'
    RP = pd.concat([正回购, 逆回购])

    RP = pd.merge(RP, get_asset('601104'), on=['账套', '科目名称'], how='left')
    RP.rename(columns={'资产代码_x': '资产代码', '买入返售债券': '净投资收益'}, inplace=True)
    RP['总投资收益'] = RP['净投资收益']
    RP['综合投资收益'] = RP['净投资收益']

    RP = RP[['账套', '资产代码', '科目名称', '资产类型描述', '币种', '数量', '成本', '资产分类', '净投资收益', '总投资收益', '综合投资收益']].fillna(0)
    RP = RP.loc[(RP['净投资收益'] != 0) | (RP['总投资收益'] != 0) | (RP['综合投资收益'] != 0)]

    # 汇总
    资产收益 = pd.concat([AFS, TS, HMS, LAR, LTR, LR, RP])
    资产收益.rename(columns={'科目名称': '资产描述'}, inplace=True)
    资产收益.loc[资产收益['账套'].str.contains('200018_财富鸿盈A|200019_财富鸿盈B|200024_财富享盈A|200025_财富享盈B'), '账套'] = '分红'
    资产收益.loc[资产收益['账套'].str.contains(
        '3003_天天盈_太保委托|2007_天天盈A-自营|3006_天天盈B-太保委托|3008_天天盈B-活动|8001_万能归集账户|200005_稳健型子账户|200004_附加天天盈B（钻石）|200009_附加天天盈B（铂金）|200008_万能归集账户2|200013_天添利年金|200012_天天盈B（紫金）|200016_天天盈年金|200017_天添利B|200022_天天盈年金（尊享版）|200026_天天盈年金（臻享版）|200027_天添利B款（尊耀）'), '账套'] = '万能'
    资产收益.loc[资产收益['账套'].str.contains(
        '1003_资本金_自营|2001_资本金-华泰委托|3007_资本金-太保委托|4003_资本金-人保委托|5004_资本金-QDII境外投资|200001_德邦-手拉手7号|200002_德邦-手拉手8号|200010_申万宏源共盈10号|200014_自营-建行托管|200020_申万宏源共盈16号|200021_浦发-自营|200007_财富固盈年金|200015_财富稳赢|200011_海通资管-01'), '账套'] = '资本金'
    资产收益 = 资产收益.groupby(['资产代码', '资产描述', '账套', '资产类型描述', '资产分类', '币种']).sum().sort_values(
        ['资产分类', '资产描述', '账套', '资产类型描述', ]).reset_index()
    资产收益 = 资产收益[['账套', '资产代码', '资产描述', '资产类型描述', '币种', '资产分类', '净投资收益', '总投资收益', '综合投资收益']]

    while True:
        # 美观输出

        table = Table(title='资产持仓查询报表', show_header=True, header_style="bold magenta")
        for i in 资产收益.columns:
            if i == '资产描述':
                table.add_column(i, style='red')
            elif i == '账套':
                table.add_column(i, style='cyan')
            else:
                table.add_column(i)
        table_centered = Align.center(table)
        with Live(table_centered, refresh_per_second=1000):
            for i in 资产收益.astype(str).values.tolist():
                time.sleep(0.001)
                table.add_row(i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8])

        Confirm = input('导出到?(excel/clipboard/no):')
        if Confirm == 'excel':
            资产收益.to_excel(Base_route + '/资产收益查询报表.xlsx', index=False)
            console.print(
                "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
            restart_program()

        elif Confirm == 'clipboard':
            资产持仓.to_clipboard(index=False)
            console.print(
                "导出成功 :smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley::smiley:")
        else:
            restart_program()
